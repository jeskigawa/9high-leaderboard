import os
import json
import uuid
import hashlib
import secrets
import csv
import io
import random
from datetime import datetime, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from functools import wraps
from flask import (
    Flask, request, jsonify, send_from_directory,
    session, abort, Response, stream_with_context, redirect
)
from werkzeug.utils import secure_filename
import requests as http_requests
from bs4 import BeautifulSoup
import re
import time
import threading
import urllib.parse

app = Flask(__name__, static_folder='static')
app.secret_key = secrets.token_hex(32)

# ==================== CONFIG ====================
DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'uploads', 'photos')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
MAX_CONTENT_LENGTH = 5 * 1024 * 1024  # 5MB

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ==================== DATA HELPERS ====================
def load_json(filename):
    path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(path):
        return {}
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_json(filename, data):
    path = os.path.join(DATA_DIR, filename)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_ringgames():
    """Load ring games, auto-migrating flat format to grouped session format.
    New format: sessions[] = [{id, name, date, pokerfansUrl, players:[{playerId, playerName, netChips}]}]
    Old format: sessions[] = [{date, playerId, playerName, netChips, pokerfansUrl?, eventName?}]
    """
    data = load_json('ringgames.json')
    if not data:
        return {'sessions': []}
    sessions = data.get('sessions', [])
    if not sessions:
        return data
    # Detect old flat format: first session lacks 'players' key
    if 'players' not in sessions[0]:
        groups = {}
        order = []
        for s in sessions:
            key = (s.get('date', ''), s.get('pokerfansUrl') or '')
            if key not in groups:
                gid = str(uuid.uuid4())
                groups[key] = {
                    'id': gid,
                    'name': s.get('eventName') or (s.get('date', '') + ' セッション'),
                    'date': s.get('date', ''),
                    'pokerfansUrl': s.get('pokerfansUrl') or None,
                    'players': []
                }
                order.append(key)
            pid = s.get('playerId')
            if pid:  # skip entries with no player ID
                groups[key]['players'].append({
                    'playerId': pid,
                    'playerName': s.get('playerName', ''),
                    'netChips': s.get('netChips', 0)
                })
        new_data = {'sessions': [groups[k] for k in order]}
        save_json('ringgames.json', new_data)
        return new_data
    return data

def load_config():
    config = load_json('config.json')
    if not config:
        config = {
            'store_name': '9HIGH 仙台',
            'store_subtitle': '月間リーダーボード',
            'users': [
                {
                    'email': 'admin@9high.jp',
                    'password_hash': hashlib.sha256('Abcd1234'.encode()).hexdigest(),
                    'role': 'admin'
                }
            ]
        }
        save_json('config.json', config)
        return config
    # Migrate old single-password config to multi-user
    if 'users' not in config:
        old_hash = config.get('password_hash', hashlib.sha256('Abcd1234'.encode()).hexdigest())
        config['users'] = [
            {'email': 'admin@9high.jp', 'password_hash': old_hash, 'role': 'admin'}
        ]
        config.pop('password_hash', None)
        save_json('config.json', config)
    return config

def load_players():
    data = load_json('players.json')
    if not data:
        data = {'players': []}
        # Migrate from existing tournament/ring data
        existing_names = set()
        tournaments = load_json('tournaments.json')
        for result in tournaments.get('results', []):
            for r in result.get('results', []):
                if r['playerId'] not in existing_names:
                    existing_names.add(r['playerId'])
                    data['players'].append({
                        'id': r['playerId'],
                        'name': r['playerName'],
                        'photo': None,
                        'nickname': '',
                        'note': '',
                        'createdAt': datetime.now().isoformat()
                    })
        ringgames = load_ringgames()
        for sess in ringgames.get('sessions', []):
            for s in sess.get('players', []):
                if s.get('playerId') and s['playerId'] not in existing_names:
                    existing_names.add(s['playerId'])
                    data['players'].append({
                        'id': s['playerId'],
                        'name': s.get('playerName', ''),
                        'photo': None,
                        'nickname': '',
                        'note': '',
                        'createdAt': datetime.now().isoformat()
                    })
        save_json('players.json', data)
    return data

# ==================== AUTH ====================
def find_user(email):
    config = load_config()
    for u in config.get('users', []):
        if u['email'].lower() == email.lower():
            return u
    return None

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('authenticated'):
            return jsonify({'error': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('authenticated'):
            return jsonify({'error': 'Unauthorized'}), 401
        if session.get('role') != 'admin':
            return jsonify({'error': '管理者権限が必要です'}), 403
        return f(*args, **kwargs)
    return decorated

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==================== ROUTES: Static ====================
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/admin')
def admin_page():
    return redirect('/', 301)

@app.route('/leaderboard')
def public_leaderboard():
    return send_from_directory('.', '9high-leaderboard.html')

@app.route('/')
def index_page():
    return send_from_directory('.', 'index.html')

@app.route('/logo.svg')
def serve_logo():
    return send_from_directory('.', 'logo.svg')

@app.route('/pokerfans')
@login_required
def pokerfans_page():
    return send_from_directory('.', 'pokerfans.html')

# ==================== ROUTES: Poker Fans ====================

import io as _io

# In-memory pokerfans session (per-process; re-login if cookie expires)
_pf_session = None
_pf_session_lock = threading.Lock()

def _get_pf_session():
    global _pf_session
    return _pf_session

def _set_pf_session(s):
    global _pf_session
    _pf_session = s

PF_BASE = 'https://pokerfans.jp'
PF_TIMEOUT = 15

def _pf_session_from_cookie(cookie_str):
    """Cookieを使ってpokerfans.jpのrequests.Sessionを作る。"""
    s = http_requests.Session()
    s.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,*/*;q=0.8',
        'Accept-Language': 'ja,en-US;q=0.9,en;q=0.8',
        'Cookie': cookie_str,
    })
    # cookieヘッダーをrequestsのcookiejarにも反映
    for part in cookie_str.split(';'):
        part = part.strip()
        if '=' in part:
            k, v = part.split('=', 1)
            s.cookies.set(k.strip(), v.strip(), domain='pokerfans.jp')
    return s

def _pf_mb_auth(s, email, password):
    """モバイルAPIセッションを取得する。成功時Trueを返す。"""
    try:
        r = s.post(PF_BASE + '/mb/auth',
                   json={'loginId': email, 'loginPw': password},
                   headers={
                       'Accept': 'application/json, text/plain, */*',
                       'Accept-Language': 'ja,en-US;q=0.9,en;q=0.8',
                       'Content-Type': 'application/json',
                       'Origin': PF_BASE,
                       'Referer': PF_BASE + '/mypage',
                       'sec-ch-ua': '"Google Chrome";v="124", "Chromium";v="124", "Not-A.Brand";v="99"',
                       'sec-ch-ua-mobile': '?0',
                       'sec-ch-ua-platform': '"Windows"',
                       'sec-fetch-dest': 'empty',
                       'sec-fetch-mode': 'cors',
                       'sec-fetch-site': 'same-origin',
                   },
                   timeout=PF_TIMEOUT)
        app.logger.info(f'[PF] mb/auth status={r.status_code} body={r.text[:300]}')
        return r.ok
    except Exception as e:
        app.logger.warning(f'[PF] mb/auth exception: {e}')
        return False

@app.route('/api/pokerfans/login', methods=['POST'])
@login_required
def pf_login():
    """CookieでPF接続を確立する。email/passwordがあれば/mb/authも実行。"""
    data = request.get_json(force=True, silent=True) or {}
    cookie_str = data.get('cookie', '').strip()
    email = data.get('email', '').strip()
    password = data.get('password', '').strip()
    if not cookie_str:
        return jsonify({'error': 'Cookieを入力してください'}), 400

    try:
        s = _pf_session_from_cookie(cookie_str)

        # 接続確認: /mypage/clubs（HTMLページ）
        r = s.get(PF_BASE + '/mypage/clubs', timeout=PF_TIMEOUT, allow_redirects=True)
        if not r.ok or '/login' in r.url:
            return jsonify({'error': 'Cookieが無効か期限切れです。ブラウザで再ログイン後に取得し直してください。'}), 400

        # モバイルAPI認証（email/passwordがある場合）
        mb_auth_ok = False
        if email and password:
            mb_auth_ok = _pf_mb_auth(s, email, password)

        # ニックネーム取得を試みる（モバイルAPIが通れば/mb/mypageから）
        nickname = ''
        try:
            mb = s.get(PF_BASE + '/mb/mypage', timeout=PF_TIMEOUT,
                       headers={'Accept': 'application/json'})
            if mb.ok:
                pd = mb.json()
                nickname = ((pd.get('clubMember') or {}).get('user') or {}).get('nickname', '')
        except Exception:
            pass
        s._pf_nickname = nickname or '接続済み'
        s._pf_mb_auth = mb_auth_ok

        with _pf_session_lock:
            _set_pf_session(s)

        # Cookieとcredentialsを保存（サーバー再起動後に自動復元）
        config = load_config()
        pf_cfg = config.setdefault('pokerfans', {})
        pf_cfg['cookie'] = cookie_str
        if email:
            pf_cfg['email'] = email
        if password:
            pf_cfg['password'] = password
        save_json('config.json', config)

        msg = s._pf_nickname
        if mb_auth_ok:
            msg += '（モバイル認証済み）'
        elif email:
            msg += '（モバイル認証失敗 — メンバー限定は取得不可）'
        return jsonify({'success': True, 'nickname': msg, 'mbAuth': mb_auth_ok})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/pokerfans/status', methods=['GET'])
@login_required
def pf_status():
    """PFセッションの状態確認。なければ保存済みCookieで自動復元を試みる。"""
    s = _get_pf_session()
    if not s:
        config = load_config()
        saved_cookie = config.get('pokerfans', {}).get('cookie', '')
        if saved_cookie:
            try:
                pf_cfg = config.get('pokerfans', {})
                ns = _pf_session_from_cookie(saved_cookie)
                r = ns.get(PF_BASE + '/mypage/clubs', timeout=PF_TIMEOUT, allow_redirects=True)
                if r.ok and '/login' not in r.url:
                    # モバイル認証（保存済みcredentials）
                    saved_email = pf_cfg.get('email', '')
                    saved_pass = pf_cfg.get('password', '')
                    if saved_email and saved_pass:
                        _pf_mb_auth(ns, saved_email, saved_pass)
                    # ニックネーム取得
                    nickname = ''
                    try:
                        mb = ns.get(PF_BASE + '/mb/mypage', timeout=PF_TIMEOUT,
                                    headers={'Accept': 'application/json'})
                        if mb.ok:
                            pd = mb.json()
                            nickname = ((pd.get('clubMember') or {}).get('user') or {}).get('nickname', '')
                    except Exception:
                        pass
                    ns._pf_nickname = nickname or '接続済み'
                    with _pf_session_lock:
                        _set_pf_session(ns)
                    return jsonify({'loggedIn': True, 'nickname': ns._pf_nickname, 'hasSavedCookie': True})
            except Exception:
                pass
        return jsonify({'loggedIn': False, 'hasSavedCookie': bool(saved_cookie)})

    # 既存セッションの有効性を確認
    try:
        r = s.get(PF_BASE + '/mypage/clubs', timeout=PF_TIMEOUT, allow_redirects=True)
        if r.ok and '/login' not in r.url:
            return jsonify({'loggedIn': True, 'nickname': getattr(s, '_pf_nickname', '接続済み')})
    except Exception:
        pass
    with _pf_session_lock:
        _set_pf_session(None)
    return jsonify({'loggedIn': False})

@app.route('/api/pokerfans/clubs', methods=['GET'])
@login_required
def pf_clubs():
    s = _get_pf_session()
    if not s:
        return jsonify({'error': 'pokerfans.jpにログインしてください'}), 401
    try:
        r = s.get(PF_BASE + '/mypage/clubs', timeout=PF_TIMEOUT)
        if not r.ok:
            return jsonify({'error': f'店舗一覧取得失敗 (HTTP {r.status_code})'}), 400
        from bs4 import BeautifulSoup as _BS
        doc = _BS(r.text, 'html.parser')
        clubs = {}
        for a in doc.find_all('a', href=True):
            m = re.search(r'/clubs/(\d+)', a['href'])
            if not m:
                continue
            cid = m.group(1)
            if cid == '1001':
                continue
            text = a.get_text(strip=True)
            if text and text not in ('See more', 'Events schedule') and not text.startswith('→') and cid not in clubs:
                clubs[cid] = text
        return jsonify({'clubs': clubs})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pokerfans/events', methods=['POST'])
@login_required
def pf_events():
    """Fetch events for selected clubs/date range. Returns SSE stream.
    Member clubs use /mb/tours/events/member (with club switching) for member-only events.
    Non-member clubs fall back to the public /events endpoint.
    """
    s = _get_pf_session()
    if not s:
        def _no_auth():
            yield f"data: {json.dumps({'type':'error','message':'pokerfans.jpにログインしてください'})}\n\n"
        return Response(stream_with_context(_no_auth()), content_type='text/event-stream')

    data = request.get_json(force=True, silent=True) or {}
    club_ids = data.get('clubIds', [])  # list of string IDs
    start_month = data.get('startMonth', '')  # YYYY-MM
    end_month = data.get('endMonth', '')      # YYYY-MM

    if not club_ids or not start_month or not end_month:
        def _bad():
            yield f"data: {json.dumps({'type':'error','message':'店舗・期間を指定してください'})}\n\n"
        return Response(stream_with_context(_bad()), content_type='text/event-stream')

    def _month_chunks(sm, em):
        import calendar
        y, mo = int(sm[:4]), int(sm[5:7])
        ey, emo = int(em[:4]), int(em[5:7])
        chunks = []
        while (y, mo) <= (ey, emo):
            last = calendar.monthrange(y, mo)[1]
            chunks.append((f'{y}/{mo:02d}/01', f'{y}/{mo:02d}/{last:02d}'))
            mo += 1
            if mo > 12:
                mo = 1
                y += 1
        return chunks

    def generate():
        all_events = []
        seen_ids = set()
        pc_map = {}  # event_id -> appliedSeats (from mobile API)

        # --- Step 1: get mypage to identify member clubs ---
        nickname = ''
        original_club_id = ''
        my_club_ids = []
        try:
            mb_r = s.get(PF_BASE + '/mb/mypage', timeout=PF_TIMEOUT,
                         headers={'Accept': 'application/json'})
            if mb_r.ok:
                pd = mb_r.json()
                user = (pd.get('clubMember') or {}).get('user') or {}
                nickname = user.get('nickname', '')
                original_club_id = str(user.get('myClubId', ''))
                raw_clubs = pd.get('myClubs') or []
                if isinstance(raw_clubs, list):
                    my_club_ids = [str(c['id']) for c in raw_clubs
                                   if str(c.get('id', '')) != '1001']
                yield f"data: {json.dumps({'type':'info','message':f'mypage OK: nickname={nickname}, myClubs={my_club_ids}'})}\n\n"
            else:
                yield f"data: {json.dumps({'type':'warning','message':f'mypage HTTP{mb_r.status_code}'})}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'type':'warning','message':f'mypage取得失敗: {e}'})}\n\n"

        sel_set = set(club_ids)
        member_targets = [cid for cid in my_club_ids if cid in sel_set]
        non_member_targets = [cid for cid in club_ids if cid not in set(my_club_ids)]
        chunks = _month_chunks(start_month, end_month)
        total = len(member_targets) + len(non_member_targets) * len(chunks)
        step = 0

        yield f"data: {json.dumps({'type':'info','message':f'MEMBER対象: {member_targets} / 公開API対象: {non_member_targets}'})}\n\n"

        # --- Step 2: member clubs via mobile API ---
        if member_targets and nickname:
            for cid in member_targets:
                step += 1
                yield f"data: {json.dumps({'type':'progress','step':step,'total':total,'clubId':cid,'period':'MEMBER'})}\n\n"

                # Switch club
                switch_ok = False
                try:
                    body = urllib.parse.urlencode({'nickname': nickname, 'myClubId': cid})
                    sr = s.put(PF_BASE + '/mb/profile/basic',
                               data=body,
                               headers={'Content-Type': 'application/x-www-form-urlencoded',
                                        'Accept': 'application/json'},
                               timeout=PF_TIMEOUT)
                    if sr.ok:
                        switch_ok = True
                        yield f"data: {json.dumps({'type':'info','message':f'店舗切替OK: {cid}'})}\n\n"
                    else:
                        yield f"data: {json.dumps({'type':'warning','message':f'店舗切替失敗 {cid}: HTTP{sr.status_code} {sr.text[:100]}'})}\n\n"
                except Exception as e:
                    yield f"data: {json.dumps({'type':'warning','message':f'店舗切替エラー {cid}: {e}'})}\n\n"

                if not switch_ok:
                    # fallback: try public API for this club
                    for (cs, ce) in chunks:
                        try:
                            r = s.get(f'{PF_BASE}/events?clubId={cid}&start={cs}&end={ce}', timeout=PF_TIMEOUT)
                            if r.ok:
                                for ev in (r.json() if isinstance(r.json(), list) else []):
                                    if ev.get('id') and ev['id'] not in seen_ids:
                                        seen_ids.add(ev['id'])
                                        ev['_clubId'] = cid
                                        all_events.append(ev)
                        except Exception:
                            pass
                    continue

                # Fetch events per month chunk from mobile member API
                for (cs, ce) in chunks:
                    try:
                        url = f'{PF_BASE}/mb/tours/events/member?start={cs}&end={ce}&size=9999'
                        r = s.get(url, timeout=PF_TIMEOUT,
                                  headers={'Accept': 'application/json'})
                        if not r.ok:
                            yield f"data: {json.dumps({'type':'warning','message':f'MEMBER API HTTP{r.status_code} {cid}/{cs[:7]}'})}\n\n"
                            continue
                        evs = r.json()
                        if not isinstance(evs, list):
                            yield f"data: {json.dumps({'type':'warning','message':f'MEMBER API 非リスト応答 {cid}/{cs[:7]}: {str(evs)[:80]}'})}\n\n"
                            continue
                        added = 0
                        for mev in evs:
                            eid = mev.get('id')
                            if not eid or eid in seen_ids:
                                continue
                            seen_ids.add(eid)
                            oid = str((mev.get('spot') or {}).get('ownerId', '') or cid)
                            ev_club = oid if oid in sel_set else cid
                            start_str = mev.get('openDate', '')
                            if mev.get('startTime'):
                                start_str += 'T' + mev['startTime']
                            end_str = ''
                            if mev.get('closeDate'):
                                end_str = mev['closeDate']
                                if mev.get('endTime'):
                                    end_str += 'T' + mev['endTime']
                            all_events.append({
                                'id': eid,
                                'title': mev.get('name', '不明'),
                                'start': start_str,
                                'end': end_str,
                                '_clubId': ev_club,
                                '_fromMobile': True,
                            })
                            if isinstance(mev.get('appliedSeats'), int):
                                pc_map[str(eid)] = mev['appliedSeats']
                            added += 1
                        yield f"data: {json.dumps({'type':'info','message':f'MEMBER {cid}/{cs[:7]}: {added}件'})}\n\n"
                    except Exception as e:
                        yield f"data: {json.dumps({'type':'warning','message':f'MEMBER API エラー {cid}/{cs[:7]}: {e}'})}\n\n"

                time.sleep(1)

            # Restore original club
            if original_club_id:
                try:
                    body = urllib.parse.urlencode({'nickname': nickname, 'myClubId': original_club_id})
                    s.put(PF_BASE + '/mb/profile/basic',
                          data=body,
                          headers={'Content-Type': 'application/x-www-form-urlencoded',
                                   'Accept': 'application/json'},
                          timeout=PF_TIMEOUT)
                except Exception:
                    pass
        elif member_targets and not nickname:
            yield f"data: {json.dumps({'type':'warning','message':'nicknameが取得できないためMEMBER APIをスキップ'})}\n\n"

        # --- Step 3: non-member clubs via public API ---
        for cid in non_member_targets:
            for (cs, ce) in chunks:
                step += 1
                yield f"data: {json.dumps({'type':'progress','step':step,'total':total,'clubId':cid,'period':cs[:7]})}\n\n"
                try:
                    url = f'{PF_BASE}/events?clubId={cid}&start={cs}&end={ce}'
                    r = s.get(url, timeout=PF_TIMEOUT)
                    if not r.ok:
                        yield f"data: {json.dumps({'type':'warning','message':f'HTTP {r.status_code}: {cid} {cs[:7]}'})}\n\n"
                        continue
                    evs = r.json()
                    if not isinstance(evs, list):
                        continue
                    for ev in evs:
                        if ev.get('id') and ev['id'] not in seen_ids:
                            seen_ids.add(ev['id'])
                            ev['_clubId'] = cid
                            all_events.append(ev)
                except Exception as ex:
                    yield f"data: {json.dumps({'type':'warning','message':str(ex)})}\n\n"

        yield f"data: {json.dumps({'type':'done','events':all_events,'count':len(all_events),'pcMap':pc_map})}\n\n"

    return Response(stream_with_context(generate()), content_type='text/event-stream')

@app.route('/api/pokerfans/participant-count', methods=['POST'])
@login_required
def pf_participant_count():
    """Fetch participant counts for a batch of event IDs. SSE stream."""
    s = _get_pf_session()
    if not s:
        def _no_auth():
            yield f"data: {json.dumps({'type':'error','message':'pokerfans.jpにログインしてください'})}\n\n"
        return Response(stream_with_context(_no_auth()), content_type='text/event-stream')

    data = request.get_json(force=True, silent=True) or {}
    event_ids = data.get('eventIds', [])
    delay_ms = max(500, int(data.get('delayMs', 2000)))

    def generate():
        # Load existing cache from server
        pf_cache = load_json('pokerfans_cache.json') or {}

        # Separate cached vs needs-fetch
        to_fetch = []
        cached_results = {}
        for eid in event_ids:
            key = str(eid)
            if key in pf_cache:
                cached_results[key] = pf_cache[key]
            else:
                to_fetch.append(eid)

        cached_count = len(cached_results)
        fetch_total = len(to_fetch)
        total = len(event_ids)

        if cached_count > 0:
            yield f"data: {json.dumps({'type':'cache','cached':cached_count,'toFetch':fetch_total})}\n\n"

        new_results = {}
        for i, eid in enumerate(to_fetch):
            url = f'{PF_BASE}/events/{eid}'
            count = 0
            try:
                r = s.get(url, timeout=PF_TIMEOUT)
                if r.ok:
                    html = r.text
                    m = re.search(r'Entries:[\s\S]*?(\d+)\s*/\s*(\d+)', html)
                    if not m:
                        m = re.search(r'人数[\/／]定員[\s\S]*?(\d+)\s*[\/／]\s*(\d+)', html)
                    if m:
                        count = int(m.group(1))
            except Exception:
                pass
            new_results[str(eid)] = count
            done_so_far = cached_count + i + 1
            yield f"data: {json.dumps({'type':'progress','done':done_so_far,'total':total,'eventId':eid,'count':count,'cached':cached_count})}\n\n"
            if i < fetch_total - 1:
                time.sleep(delay_ms / 1000)

        # Merge and save cache
        pf_cache.update(new_results)
        save_json('pokerfans_cache.json', pf_cache)

        all_results = {**cached_results, **new_results}
        yield f"data: {json.dumps({'type':'done','results':all_results,'newFetched':fetch_total,'fromCache':cached_count})}\n\n"

    return Response(stream_with_context(generate()), content_type='text/event-stream')

def _cors_upload_response(resp):
    """ブックマークレット(pokerfans.jp)からのCORSを許可"""
    origin = request.headers.get('Origin', '')
    if 'pokerfans.jp' in origin:
        resp.headers['Access-Control-Allow-Origin'] = origin
    else:
        resp.headers['Access-Control-Allow-Origin'] = 'https://pokerfans.jp'
    resp.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
    resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    resp.headers['Access-Control-Allow-Credentials'] = 'true'
    return resp

@app.route('/api/pokerfans/upload', methods=['POST', 'OPTIONS'])
def pf_upload():
    """ブックマークレットからデータを受け取りサーバーに保存する"""
    # OPTIONSはpreflight — 認証不要で即返す
    if request.method == 'OPTIONS':
        return _cors_upload_response(Response('', 204))

    # POST は通常ログイン必須だが、ブックマークレットはセッションCookieを持てないため
    # 代わりにローカルからのアクセスのみ受け付ける（127.0.0.1 / ::1）
    remote = request.remote_addr
    if remote not in ('127.0.0.1', '::1', 'localhost'):
        return _cors_upload_response(jsonify({'error': 'ローカルからのアクセスのみ許可されています'})), 403

    data = request.get_json(force=True, silent=True) or {}
    clubs      = data.get('clubs', {})
    result_data = data.get('resultData', {})
    date_list  = data.get('dateList', [])
    start_month = data.get('startMonth', '')
    end_month   = data.get('endMonth', '')

    if not clubs or not result_data:
        return _cors_upload_response(jsonify({'error': 'データが空です'})), 400

    payload = {
        'clubs': clubs,
        'resultData': result_data,
        'dateList': date_list,
        'startMonth': start_month,
        'endMonth': end_month,
        'uploadedAt': time.strftime('%Y-%m-%d %H:%M:%S'),
    }
    save_json('pokerfans_data.json', payload)
    total_events = sum(
        len(evs)
        for club_data in result_data.values()
        for evs in club_data.values()
    )
    resp = jsonify({'success': True, 'clubs': len(clubs), 'events': total_events})
    return _cors_upload_response(resp)

@app.route('/api/pokerfans/data', methods=['GET'])
@login_required
def pf_data():
    """保存済みデータを返す"""
    data = load_json('pokerfans_data.json')
    if not data:
        return jsonify({'error': 'データがありません。ブックマークレットからアップロードしてください。'}), 404
    return jsonify(data)

@app.route('/api/pokerfans/cache', methods=['GET'])
@login_required
def pf_cache_info():
    """キャッシュの状態を返す"""
    cache = load_json('pokerfans_cache.json') or {}
    return jsonify({'count': len(cache)})

@app.route('/api/pokerfans/cache', methods=['DELETE'])
@login_required
def pf_cache_clear():
    """キャッシュをリセット"""
    save_json('pokerfans_cache.json', {})
    return jsonify({'success': True})

@app.route('/api/pokerfans/report', methods=['POST'])
@login_required
def pf_report():
    """Generate Excel report from provided data."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = request.get_json(force=True, silent=True) or {}
    clubs = data.get('clubs', {})       # {clubId: name}
    result_data = data.get('resultData', {})  # {clubId: {date: [{title, count}]}}
    date_list = data.get('dateList', [])  # sorted list of YYYY-MM-DD strings
    start_month = data.get('startMonth', '')
    end_month = data.get('endMonth', '')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    HDR_FILL = PatternFill('solid', fgColor='2BBBAD')
    HDR_FONT = Font(color='FFFFFF', bold=True, size=9)
    TOTAL_FILL = PatternFill('solid', fgColor='FFF3E0')
    TOTAL_FONT = Font(bold=True, size=9)
    WE_FILL = PatternFill('solid', fgColor='E3F2FD')
    SUMMARY_HDR_FILL = PatternFill('solid', fgColor='607D8B')
    thin = Side(style='thin', color='BDBDBD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    from datetime import datetime as _dt
    def is_weekend(d):
        try:
            return _dt.strptime(d, '%Y-%m-%d').weekday() >= 5
        except Exception:
            return False

    def fmt_date(d):
        try:
            dt = _dt.strptime(d, '%Y-%m-%d')
            wd = ['月','火','水','木','金','土','日'][dt.weekday()]
            return f"{dt.month}/{dt.day}({wd})"
        except Exception:
            return d

    for cid, club_name in clubs.items():
        if cid not in result_data:
            continue
        sd = result_data[cid]
        max_ev = max((len(sd.get(d, [])) for d in date_list), default=0)
        if max_ev == 0:
            continue

        sheet_name = re.sub(r'[\[\]*?\/\\:]', '', club_name)[:31] or cid
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        headers = ['日付', '合計']
        for i in range(max_ev):
            headers += [f'イベント{i+1}', f'人数{i+1}']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = center
            cell.border = border

        # Data rows
        summary = {}
        for row_i, d in enumerate(date_list, 2):
            evs = sd.get(d, [])
            we = is_weekend(d)
            day_total = sum(e.get('count', 0) for e in evs)

            date_cell = ws.cell(row=row_i, column=1, value=fmt_date(d))
            date_cell.alignment = left
            date_cell.border = border
            if we:
                date_cell.fill = WE_FILL

            total_cell = ws.cell(row=row_i, column=2, value=day_total if day_total > 0 else '')
            total_cell.fill = TOTAL_FILL if not we else WE_FILL
            total_cell.font = TOTAL_FONT
            total_cell.alignment = center
            total_cell.border = border

            for ei, ev in enumerate(evs):
                tc = ws.cell(row=row_i, column=3 + ei*2, value=ev.get('title', ''))
                tc.alignment = left
                tc.border = border
                if we:
                    tc.fill = WE_FILL

                nc = ws.cell(row=row_i, column=4 + ei*2, value=ev.get('count', 0) or '')
                nc.alignment = center
                nc.border = border
                if we:
                    nc.fill = WE_FILL

                title = ev.get('title', '不明')
                cnt = ev.get('count', 0) or 0
                if title not in summary:
                    summary[title] = {'scheduled': 0, 'actual': 0, 'total': 0}
                summary[title]['scheduled'] += 1
                if cnt > 0:
                    summary[title]['actual'] += 1
                    summary[title]['total'] += cnt

            # Empty padding cells
            for pad in range(len(evs), max_ev):
                for col_off in [0, 1]:
                    c = ws.cell(row=row_i, column=3 + pad*2 + col_off, value='')
                    c.border = border

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        for i in range(max_ev):
            ws.column_dimensions[get_column_letter(3 + i*2)].width = 20
            ws.column_dimensions[get_column_letter(4 + i*2)].width = 7

        # Summary table (below data, 2 blank rows)
        sum_start = len(date_list) + 4
        sum_headers = ['イベント名', '予定回数', '実施回数', '実施割合', '参加総数', '平均参加']
        for col, h in enumerate(sum_headers, 1):
            cell = ws.cell(row=sum_start, column=col, value=h)
            cell.fill = SUMMARY_HDR_FILL
            cell.font = Font(color='FFFFFF', bold=True, size=9)
            cell.alignment = center
            cell.border = border

        for si, (title, s2) in enumerate(sorted(summary.items()), 1):
            pct = round(s2['actual'] / s2['scheduled'] * 100, 1) if s2['scheduled'] > 0 else 0
            avg = round(s2['total'] / s2['actual'], 1) if s2['actual'] > 0 else 0
            row_vals = [title, s2['scheduled'], s2['actual'], f"{pct}%", s2['total'], avg]
            for col, v in enumerate(row_vals, 1):
                c = ws.cell(row=sum_start + si, column=col, value=v)
                c.alignment = left if col == 1 else center
                c.border = border

    if not wb.sheetnames:
        ws = wb.create_sheet('データなし')
        ws['A1'] = 'データがありません'

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"pokerfans_{start_month}_{end_month}.xlsx".replace(':', '-')
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'}
    )

@app.route('/uploads/photos/<path:filename>')
def uploaded_photo(filename):
    return send_from_directory(UPLOAD_DIR, filename)

# ==================== ROUTES: Auth ====================
@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json(force=True, silent=True) or {}
    email = (data.get('email') or '').strip()
    password = data.get('password', '')
    if not email or not password:
        return jsonify({'error': 'メールアドレスとパスワードを入力してください'}), 400
    user = find_user(email)
    if not user:
        return jsonify({'error': 'メールアドレスまたはパスワードが正しくありません'}), 401
    if hashlib.sha256(password.encode()).hexdigest() != user['password_hash']:
        return jsonify({'error': 'メールアドレスまたはパスワードが正しくありません'}), 401
    session['authenticated'] = True
    session['email'] = user['email']
    session['role'] = user['role']
    return jsonify({'success': True, 'role': user['role'], 'email': user['email']})

def send_reset_email(to_email, reset_url):
    config = load_config()
    smtp = config.get('smtp', {})
    if not smtp.get('host') or not smtp.get('user') or not smtp.get('password'):
        raise Exception('メール設定が未完了です。管理者にお問い合わせください。')
    msg = MIMEMultipart()
    msg['From'] = smtp.get('from_addr') or smtp['user']
    msg['To'] = to_email
    msg['Subject'] = '【9HIGH】パスワードリセット'
    body = f"""パスワードリセットのリクエストを受け付けました。

以下のリンクをクリックしてパスワードを再設定してください（有効期限: 1時間）:

{reset_url}

このメールに心当たりがない場合は無視してください。
"""
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    with smtplib.SMTP(smtp['host'], int(smtp.get('port', 587))) as server:
        server.starttls()
        server.login(smtp['user'], smtp['password'])
        server.send_message(msg)

@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    data = request.get_json(force=True, silent=True) or {}
    email = (data.get('email') or '').strip().lower()
    user = find_user(email)
    if user:
        config = load_config()
        token = secrets.token_urlsafe(32)
        expires = (datetime.now() + timedelta(hours=1)).isoformat()
        tokens = [t for t in config.get('reset_tokens', []) if t['email'].lower() != email]
        tokens.append({'email': email, 'token': token, 'expires': expires})
        config['reset_tokens'] = tokens
        save_json('config.json', config)
        base_url = config.get('smtp', {}).get('base_url', 'http://localhost:5000')
        reset_url = f"{base_url}/admin?reset_token={token}"
        try:
            send_reset_email(email, reset_url)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    # Always return success to prevent email enumeration
    return jsonify({'success': True})

@app.route('/api/reset-password', methods=['POST'])
def reset_password_route():
    data = request.get_json(force=True, silent=True) or {}
    token = data.get('token', '')
    new_password = data.get('password', '')
    if not token or not new_password:
        return jsonify({'error': 'トークンとパスワードが必要です'}), 400
    if len(new_password) < 6:
        return jsonify({'error': 'パスワードは6文字以上にしてください'}), 400
    config = load_config()
    token_entry = next((t for t in config.get('reset_tokens', []) if t['token'] == token), None)
    if not token_entry:
        return jsonify({'error': 'リンクが無効または使用済みです'}), 400
    if datetime.fromisoformat(token_entry['expires']) < datetime.now():
        return jsonify({'error': 'リンクの有効期限が切れています（1時間）'}), 400
    email = token_entry['email']
    for u in config.get('users', []):
        if u['email'].lower() == email.lower():
            u['password_hash'] = hashlib.sha256(new_password.encode()).hexdigest()
            break
    config['reset_tokens'] = [t for t in config.get('reset_tokens', []) if t['token'] != token]
    save_json('config.json', config)
    return jsonify({'success': True})

@app.route('/api/smtp-settings', methods=['GET'])
@admin_required
def get_smtp_settings():
    config = load_config()
    smtp = config.get('smtp', {})
    return jsonify({
        'host': smtp.get('host', ''),
        'port': smtp.get('port', 587),
        'user': smtp.get('user', ''),
        'from_addr': smtp.get('from_addr', ''),
        'base_url': smtp.get('base_url', 'http://localhost:5000'),
        'configured': bool(smtp.get('host') and smtp.get('user') and smtp.get('password'))
    })

@app.route('/api/smtp-settings', methods=['POST'])
@admin_required
def save_smtp_settings():
    data = request.get_json(force=True, silent=True) or {}
    config = load_config()
    smtp = config.get('smtp', {})
    for key in ('host', 'port', 'user', 'password', 'from_addr', 'base_url'):
        if data.get(key) is not None and data[key] != '':
            smtp[key] = data[key]
    config['smtp'] = smtp
    save_json('config.json', config)
    return jsonify({'success': True})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/auth/check', methods=['GET'])
def auth_check():
    return jsonify({
        'authenticated': session.get('authenticated', False),
        'role': session.get('role', ''),
        'email': session.get('email', '')
    })

# ==================== ROUTES: Users (admin only) ====================
@app.route('/api/users', methods=['GET'])
@admin_required
def list_users():
    config = load_config()
    users = [{'email': u['email'], 'role': u['role']} for u in config.get('users', [])]
    return jsonify({'users': users})

@app.route('/api/users', methods=['POST'])
@admin_required
def add_user():
    data = request.get_json()
    email = (data.get('email') or '').strip()
    password = data.get('password', '')
    role = data.get('role', 'user')
    if not email or not password:
        return jsonify({'error': 'メールアドレスとパスワードを入力してください'}), 400
    if role not in ('admin', 'user'):
        role = 'user'
    if len(password) < 6:
        return jsonify({'error': 'パスワードは6文字以上にしてください'}), 400
    config = load_config()
    if any(u['email'].lower() == email.lower() for u in config.get('users', [])):
        return jsonify({'error': 'そのメールアドレスは既に登録されています'}), 400
    config.setdefault('users', []).append({
        'email': email,
        'password_hash': hashlib.sha256(password.encode()).hexdigest(),
        'role': role
    })
    save_json('config.json', config)
    return jsonify({'success': True})

@app.route('/api/users/<path:email>', methods=['DELETE'])
@admin_required
def delete_user(email):
    if email.lower() == session.get('email', '').lower():
        return jsonify({'error': '自分自身は削除できません'}), 400
    config = load_config()
    before = len(config.get('users', []))
    config['users'] = [u for u in config.get('users', []) if u['email'].lower() != email.lower()]
    if len(config['users']) == before:
        return jsonify({'error': 'ユーザーが見つかりません'}), 404
    save_json('config.json', config)
    return jsonify({'success': True})

@app.route('/api/users/change-password', methods=['POST'])
@login_required
def change_own_password():
    data = request.get_json()
    current = data.get('currentPassword', '')
    new_pw = data.get('newPassword', '')
    if len(new_pw) < 6:
        return jsonify({'error': 'パスワードは6文字以上にしてください'}), 400
    config = load_config()
    user = find_user(session.get('email', ''))
    if not user:
        return jsonify({'error': 'ユーザーが見つかりません'}), 404
    if hashlib.sha256(current.encode()).hexdigest() != user['password_hash']:
        return jsonify({'error': '現在のパスワードが正しくありません'}), 400
    user['password_hash'] = hashlib.sha256(new_pw.encode()).hexdigest()
    save_json('config.json', config)
    return jsonify({'success': True})

# ==================== ROUTES: Config ====================
@app.route('/api/config', methods=['GET'])
def get_config():
    config = load_config()
    return jsonify({
        'store_name': config.get('store_name', '9HIGH 仙台'),
        'store_subtitle': config.get('store_subtitle', '月間リーダーボード')
    })

@app.route('/api/config', methods=['PUT'])
@login_required
def update_config():
    data = request.get_json()
    config = load_config()
    config['store_name'] = data.get('store_name', config['store_name'])
    config['store_subtitle'] = data.get('store_subtitle', config['store_subtitle'])
    save_json('config.json', config)
    return jsonify({'success': True})

# ==================== ROUTES: Players ====================
@app.route('/api/players', methods=['GET'])
def get_players():
    data = load_players()
    return jsonify(data)

@app.route('/api/players', methods=['POST'])
@login_required
def create_player():
    req = request.get_json()
    data = load_players()
    player = {
        'id': 'p' + str(uuid.uuid4())[:8],
        'name': req.get('name', ''),
        'photo': None,
        'photoUrl': req.get('photoUrl', ''),
        'nickname': req.get('nickname', ''),
        'memberCode': req.get('memberCode', ''),
        'note': req.get('note', ''),
        'createdAt': datetime.now().isoformat()
    }
    data['players'].append(player)
    save_json('players.json', data)
    return jsonify(player), 201

@app.route('/api/players/<player_id>', methods=['PUT'])
@login_required
def update_player(player_id):
    req = request.get_json()
    data = load_players()
    for p in data['players']:
        if p['id'] == player_id:
            p['name'] = req.get('name', p['name'])
            p['nickname'] = req.get('nickname', p.get('nickname', ''))
            p['note'] = req.get('note', p.get('note', ''))
            if 'photoUrl' in req:
                p['photoUrl'] = req['photoUrl']
            if 'memberCode' in req:
                p['memberCode'] = req['memberCode']
            save_json('players.json', data)
            # Also update name in tournaments and ring games
            _update_player_name_everywhere(player_id, p['name'])
            return jsonify(p)
    return jsonify({'error': 'Player not found'}), 404

@app.route('/api/players/<player_id>/photo', methods=['POST'])
@login_required
def upload_player_photo(player_id):
    if 'photo' not in request.files:
        return jsonify({'error': 'ファイルが選択されていません'}), 400
    file = request.files['photo']
    if file.filename == '':
        return jsonify({'error': 'ファイルが選択されていません'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': '対応していないファイル形式です（PNG, JPG, GIF, WebP）'}), 400

    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f"{player_id}_{uuid.uuid4().hex[:8]}.{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    file.save(filepath)

    data = load_players()
    for p in data['players']:
        if p['id'] == player_id:
            # Delete old photo
            if p.get('photo'):
                old_path = os.path.join(UPLOAD_DIR, p['photo'])
                if os.path.exists(old_path):
                    os.remove(old_path)
            p['photo'] = filename
            save_json('players.json', data)
            return jsonify({'photo': filename, 'url': f'/uploads/photos/{filename}'})
    return jsonify({'error': 'Player not found'}), 404

@app.route('/api/players/<player_id>', methods=['DELETE'])
@login_required
def delete_player(player_id):
    data = load_players()
    data['players'] = [p for p in data['players'] if p['id'] != player_id]
    save_json('players.json', data)
    return jsonify({'success': True})

def _update_player_name_everywhere(player_id, new_name):
    # Update tournaments
    tournaments = load_json('tournaments.json')
    for event in tournaments.get('results', []):
        for r in event.get('results', []):
            if r['playerId'] == player_id:
                r['playerName'] = new_name
    save_json('tournaments.json', tournaments)
    # Update ring games (grouped format: sessions[].players[])
    ringgames = load_ringgames()
    for sess in ringgames.get('sessions', []):
        for p in sess.get('players', []):
            if p.get('playerId') == player_id:
                p['playerName'] = new_name
    save_json('ringgames.json', ringgames)

# ==================== ROUTES: Tournaments ====================
@app.route('/api/tournaments', methods=['GET'])
def get_tournaments():
    data = load_json('tournaments.json')
    return jsonify(data)

@app.route('/api/tournaments/types', methods=['GET'])
def get_tournament_types():
    data = load_json('tournaments.json')
    return jsonify(data.get('tournaments', []))

@app.route('/api/tournaments/types', methods=['POST'])
@login_required
def create_tournament_type():
    req = request.get_json()
    data = load_json('tournaments.json')
    if 'tournaments' not in data:
        data['tournaments'] = []
    t = {
        'id': req.get('id', 'tourney-' + uuid.uuid4().hex[:6]),
        'name': req.get('name', ''),
        'type': req.get('type', 'NLH')
    }
    data['tournaments'].append(t)
    save_json('tournaments.json', data)
    return jsonify(t), 201

@app.route('/api/tournaments/results', methods=['POST'])
@login_required
def add_tournament_result():
    req = request.get_json()
    data = load_json('tournaments.json')
    if 'results' not in data:
        data['results'] = []
    event = {
        'tournamentId': req.get('tournamentId'),
        'date': req.get('date'),
        'entries': req.get('entries', 0),
        'results': req.get('results', [])
    }
    data['results'].append(event)
    save_json('tournaments.json', data)
    return jsonify(event), 201

@app.route('/api/tournaments/results/<int:idx>', methods=['PUT'])
@login_required
def update_tournament_result(idx):
    req = request.get_json()
    data = load_json('tournaments.json')
    results = data.get('results', [])
    if 0 <= idx < len(results):
        results[idx] = {
            'tournamentId': req.get('tournamentId'),
            'date': req.get('date'),
            'entries': req.get('entries', 0),
            'results': req.get('results', [])
        }
        save_json('tournaments.json', data)
        return jsonify(results[idx])
    return jsonify({'error': 'Not found'}), 404

@app.route('/api/tournaments/results/<int:idx>', methods=['DELETE'])
@login_required
def delete_tournament_result(idx):
    data = load_json('tournaments.json')
    results = data.get('results', [])
    if 0 <= idx < len(results):
        results.pop(idx)
        save_json('tournaments.json', data)
        return jsonify({'success': True})
    return jsonify({'error': 'Not found'}), 404

# ==================== ROUTES: Ring Games ====================
@app.route('/api/ringgames', methods=['GET'])
def get_ringgames():
    data = load_ringgames()
    return jsonify(data)

@app.route('/api/ringgames/sessions', methods=['POST'])
@login_required
def add_ring_session():
    """Create a new ring game session group.
    Body: { name, date, players: [{playerId, netChips}] }
    """
    req = request.get_json()
    data = load_ringgames()
    players_data = load_players()
    player_map = {p['id']: p['name'] for p in players_data.get('players', [])}

    players = []
    for entry in req.get('players', []):
        pid = entry.get('playerId')
        pname = player_map.get(pid, entry.get('playerName', '')) if pid else entry.get('playerName', '')
        players.append({
            'playerId': pid or None,
            'playerName': pname,
            'netChips': entry.get('netChips', 0)
        })

    date_val = req.get('date', '')
    new_session = {
        'id': str(uuid.uuid4()),
        'name': req.get('name') or (date_val + ' セッション'),
        'date': date_val,
        'pokerfansUrl': None,
        'players': players
    }
    data['sessions'].append(new_session)
    save_json('ringgames.json', data)
    return jsonify(new_session), 201

@app.route('/api/ringgames/sessions/<session_id>', methods=['PUT'])
@login_required
def update_ring_session(session_id):
    """Update a ring game session group by UUID.
    Body: { name?, date?, players?: [{playerId, netChips}] }
    """
    req = request.get_json()
    data = load_ringgames()
    sess = next((s for s in data.get('sessions', []) if s.get('id') == session_id), None)
    if not sess:
        return jsonify({'error': 'セッションが見つかりません'}), 404

    if 'name' in req:
        sess['name'] = req['name']
    if 'date' in req:
        sess['date'] = req['date']
    if 'players' in req:
        players_data = load_players()
        player_map = {p['id']: p['name'] for p in players_data.get('players', [])}
        new_players = []
        for entry in req['players']:
            pid = entry.get('playerId')
            pname = player_map.get(pid, entry.get('playerName', '')) if pid else entry.get('playerName', '')
            new_players.append({
                'playerId': pid or None,
                'playerName': pname,
                'netChips': entry.get('netChips', 0)
            })
        sess['players'] = new_players

    save_json('ringgames.json', data)
    return jsonify(sess)

@app.route('/api/ringgames/sessions/<session_id>', methods=['DELETE'])
@login_required
def delete_ring_session(session_id):
    """Delete a ring game session group by UUID."""
    data = load_ringgames()
    before = len(data.get('sessions', []))
    data['sessions'] = [s for s in data.get('sessions', []) if s.get('id') != session_id]
    if len(data['sessions']) == before:
        return jsonify({'error': 'セッションが見つかりません'}), 404
    save_json('ringgames.json', data)
    return jsonify({'success': True})

# ==================== DATA HELPER: Prizes ====================
def load_prizes():
    data = load_json('prizes.json')
    if not data:
        data = {'types': [], 'awards': []}
        save_json('prizes.json', data)
    return data

# ==================== ROUTES: Sunvy Import ====================

# --- Safety: Global rate & concurrency controls for Sunvy access ---
SUNVY_LOCK = threading.Lock()          # Prevent concurrent fetches
SUNVY_LAST_FETCH_TIME = 0              # Epoch of last completed fetch
SUNVY_MIN_FETCH_INTERVAL = 3600        # Minimum 1 hour between full fetches
SUNVY_MAX_PAGES = 50                   # Hard cap on page count (safety net)
SUNVY_PAGE_INTERVAL = 10               # Seconds between page requests
SUNVY_MAX_CONSECUTIVE_ERRORS = 3       # Stop after N consecutive failures
SUNVY_REQUEST_TIMEOUT = 15             # Timeout per HTTP request (seconds)
SUNVY_EVENT_LOCK = threading.Lock()    # Prevent concurrent event imports

# Background import status — shared between generator thread and status endpoint
_import_log = []          # recent log messages: [{type, message, time}]
_import_log_lock = threading.Lock()
MAX_IMPORT_LOG = 150      # keep last N messages

def _append_import_log(msg_type, message):
    with _import_log_lock:
        _import_log.append({
            'type': msg_type,
            'message': message,
            'time': datetime.now().isoformat(),
        })
        if len(_import_log) > MAX_IMPORT_LOG:
            del _import_log[0]

def _clear_import_log():
    with _import_log_lock:
        _import_log.clear()

def parse_sunvy_nickname(raw_nickname):
    """Parse 'XXX(YYYYYY)' or 'XXX(YYYYYY)--' etc. into name and member code."""
    # Allow trailing characters like '--' after the parentheses
    match = re.match(r'^(.+?)\(([^)]+)\)', raw_nickname.strip())
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return raw_nickname.strip(), ''

def resolve_photo_url(src):
    if not src:
        return ''
    if src.startswith('//'):
        return 'https:' + src
    if src.startswith('/'):
        return 'https://club.sunvy.jp' + src
    if src.startswith('http'):
        return src
    return ''

def parse_members_from_page(page_soup):
    """Parse member rows from a Sunvy members page table."""
    members = []
    table_rows = page_soup.find_all('tr')
    for row in table_rows:
        tds = row.find_all('td')
        if len(tds) < 2:
            continue

        # Column 1: Picture (img in first td)
        img = tds[0].find('img')
        photo_url = resolve_photo_url(img.get('src', '') if img else '')

        # Column 2: Nickname like "名前(123456)"
        nickname_td = tds[1]
        # Get the link text or first text node (nickname is usually in an <a> tag)
        nickname_link = nickname_td.find('a')
        raw = nickname_link.get_text(strip=True) if nickname_link else nickname_td.get_text(strip=True)
        if not raw:
            continue

        name, member_code = parse_sunvy_nickname(raw)
        if not name:
            continue

        # 不正行フィルター: ページ番号・日付・極端に短い文字列を除外
        if re.match(r'^\d+$', name):          # 数値のみ（例: "20"）
            continue
        if re.match(r'^\d{1,2}/\d{1,2}$', name):  # MM/DD形式（例: "2/15"）
            continue
        if len(name) < 2:                      # 1文字以下
            continue
        # メンバーコードがあれば6桁の数字であるはず
        if member_code and not re.match(r'^\d{4,8}$', member_code):
            continue

        # Column 3: "9Highポイント / 9Highドル\nMM/DD テキスト"
        # e.g. "86127 / 133210\n06/28 リエントリー"
        points = ''
        dollars = ''
        last_visit = ''

        if len(tds) > 3:
            td3_text = tds[3].get_text(separator='\n', strip=True)
            for line in td3_text.split('\n'):
                line = line.strip()
                if not line or line in ('-', '—', '－', '/'):
                    continue
                # "POINTS / DOLLARS" pattern (e.g. "86127 / 133210")
                slash_match = re.match(r'^\s*([\d,]+)\s*/\s*([\d,]+)\s*$', line)
                if slash_match:
                    points = slash_match.group(1).replace(',', '')
                    dollars = slash_match.group(2).replace(',', '')
                    continue
                # Recent Visit: MM/DD format (e.g. "06/28 リエントリー")
                if not last_visit:
                    date_match = re.search(r'(?<!\d)(\d{1,2})/(\d{1,2})(?!\d)', line)
                    if date_match:
                        mm = date_match.group(1).zfill(2)
                        dd = date_match.group(2).zfill(2)
                        last_visit = f"{mm}/{dd}"

        members.append({
            'name': name,
            'memberCode': member_code,
            'photoUrl': photo_url,
            'raw': raw,
            'points': points,
            'dollars': dollars,
            'lastVisit': last_visit
        })
    return members

def get_max_page(page_soup):
    """Find the last page number from pagination links."""
    # Look for pagination: <ul class="pagination"> with <li><a href="?page=N">
    pagination = page_soup.find('ul', class_='pagination')
    if not pagination:
        return 1
    max_page = 1
    for link in pagination.find_all('a'):
        href = link.get('href', '')
        page_match = re.search(r'[?&]page=(\d+)', href)
        if page_match:
            p = int(page_match.group(1))
            if p > max_page:
                max_page = p
    return max_page


def sunvy_login(email, password):
    """Log into club.sunvy.jp with email/password. Returns authenticated requests.Session."""
    s = http_requests.Session()
    s.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})

    login_url = 'https://club.sunvy.jp/login'
    try:
        login_page = s.get(login_url, timeout=SUNVY_REQUEST_TIMEOUT)
    except Exception:
        raise Exception('Sunvyサーバーに接続できません')

    # Extract Spring Security CSRF token if present
    soup = BeautifulSoup(login_page.text, 'html.parser')
    csrf_input = soup.find('input', {'name': '_csrf'})
    csrf_token = csrf_input.get('value', '') if csrf_input else ''

    form_data = {'username': email, 'password': password}
    if csrf_token:
        form_data['_csrf'] = csrf_token

    resp = s.post(login_url, data=form_data, timeout=SUNVY_REQUEST_TIMEOUT, allow_redirects=True)

    # Login failed if redirected back to login page
    if '/login' in resp.url:
        raise Exception('ログインに失敗しました。メールアドレスまたはパスワードを確認してください。')

    return s


@app.route('/api/sunvy/credentials', methods=['GET'])
@login_required
def get_sunvy_credentials():
    config = load_config()
    sc = config.get('sunvy_creds', {})
    return jsonify({'email': sc.get('email', ''), 'hasPassword': bool(sc.get('password', ''))})

@app.route('/api/sunvy/credentials', methods=['POST'])
@login_required
def save_sunvy_credentials():
    req = request.get_json()
    email = req.get('email', '').strip()
    password = req.get('password', '').strip()
    config = load_config()
    sc = config.setdefault('sunvy_creds', {})
    if email:
        sc['email'] = email
    if password:
        sc['password'] = password
    save_json('config.json', config)
    return jsonify({'success': True})


@app.route('/api/sunvy/fetch-members', methods=['POST'])
@login_required
def sunvy_fetch_members():
    global SUNVY_LAST_FETCH_TIME

    req = request.get_json()
    session_cookie = req.get('sessionCookie', '').strip()
    use_cache = req.get('useCache', False)

    # --- Safety 1: Use cache if available and requested ---
    cache = load_json('sunvy_cache.json')
    if use_cache and cache.get('members') and cache.get('cachedAt'):
        return jsonify({
            'members': cache['members'],
            'count': len(cache['members']),
            'cached': True,
            'cachedAt': cache['cachedAt']
        })

    if not session_cookie:
        return jsonify({'error': 'セッションCookieを入力してください'}), 400

    # --- Safety 2: Minimum interval between full fetches (1 hour) ---
    now = time.time()
    elapsed = now - SUNVY_LAST_FETCH_TIME
    if elapsed < SUNVY_MIN_FETCH_INTERVAL:
        remaining_min = int((SUNVY_MIN_FETCH_INTERVAL - elapsed) / 60)
        if cache.get('members') and cache.get('cachedAt'):
            return jsonify({
                'members': cache['members'],
                'count': len(cache['members']),
                'cached': True,
                'cachedAt': cache['cachedAt'],
                'rateLimited': True,
                'message': f'サーバー保護のため次回取得まで約{remaining_min}分お待ちください。キャッシュを表示しています。'
            })
        return jsonify({
            'error': f'サーバー保護のため次回取得まで約{remaining_min}分お待ちください'
        }), 429

    # --- Safety 3: Prevent concurrent fetches ---
    if not SUNVY_LOCK.acquire(blocking=False):
        return jsonify({'error': '別の取得処理が実行中です。完了までお待ちください。'}), 409

    def generate():
        global SUNVY_LAST_FETCH_TIME
        try:
            s = http_requests.Session()
            s.cookies.set('JSESSIONID', session_cookie, domain='club.sunvy.jp')

            all_members = []
            consecutive_errors = 0
            page = 0
            seen_keys = set()  # For duplicate/wrap-around detection

            while page <= SUNVY_MAX_PAGES:
                if page == 1:
                    yield f"data: {json.dumps({'type': 'status', 'message': 'メンバー一覧を取得中...'})}\n\n"
                else:
                    wait_msg = f'次のページまで{SUNVY_PAGE_INTERVAL}秒待機中... (ページ{page})'
                    yield f"data: {json.dumps({'type': 'waiting', 'page': page, 'message': wait_msg})}\n\n"
                    time.sleep(SUNVY_PAGE_INTERVAL)

                try:
                    page_url = f'https://club.sunvy.jp/members?keyword=&size=100&page={page}'
                    page_resp = s.get(page_url, timeout=SUNVY_REQUEST_TIMEOUT)

                    # Check if redirected to login (session expired)
                    if '/login' in page_resp.url:
                        yield f"data: {json.dumps({'type': 'error', 'message': 'セッションが無効です。ブラウザでSunvyに再ログインしてCookieを取得し直してください。'})}\n\n"
                        return

                    if page_resp.status_code == 429:
                        yield f"data: {json.dumps({'type': 'warning', 'message': 'Sunvyサーバーからレート制限を受けました。取得済みデータを保存して停止します。'})}\n\n"
                        break
                    if page_resp.status_code >= 500:
                        consecutive_errors += 1
                        yield f"data: {json.dumps({'type': 'warning', 'message': f'ページ {page}: サーバーエラー（{page_resp.status_code}）'})}\n\n"
                        if consecutive_errors >= SUNVY_MAX_CONSECUTIVE_ERRORS:
                            yield f"data: {json.dumps({'type': 'warning', 'message': f'{SUNVY_MAX_CONSECUTIVE_ERRORS}回連続エラーのため停止します。取得済みデータを保存します。'})}\n\n"
                            break
                        time.sleep(SUNVY_PAGE_INTERVAL)
                        page += 1
                        continue
                    if page_resp.status_code >= 400:
                        consecutive_errors += 1
                        yield f"data: {json.dumps({'type': 'warning', 'message': f'ページ {page}: エラー（{page_resp.status_code}）'})}\n\n"
                        if consecutive_errors >= SUNVY_MAX_CONSECUTIVE_ERRORS:
                            yield f"data: {json.dumps({'type': 'warning', 'message': f'{SUNVY_MAX_CONSECUTIVE_ERRORS}回連続エラーのため停止します。'})}\n\n"
                            break
                        page += 1
                        continue

                    consecutive_errors = 0
                    page_html = page_resp.text
                    page_soup = BeautifulSoup(page_html, 'html.parser')

                    # Stop if page contains empty-state text (Japanese or English)
                    page_text = page_soup.get_text()
                    if 'データがありません' in page_text or 'No Data' in page_text:
                        yield f"data: {json.dumps({'type': 'status', 'message': f'ページ{page}: データなし。全ページ取得完了。'})}\n\n"
                        break

                    page_members = parse_members_from_page(page_soup)

                    # Stop if no members parsed
                    if len(page_members) == 0:
                        yield f"data: {json.dumps({'type': 'status', 'message': f'ページ{page}: メンバーなし。全ページ取得完了。'})}\n\n"
                        break

                    # Duplicate/wrap-around detection: stop if all members already seen
                    page_keys = {m.get('memberCode') or m.get('name', '') for m in page_members}
                    new_keys = page_keys - seen_keys
                    if len(new_keys) == 0:
                        yield f"data: {json.dumps({'type': 'status', 'message': f'ページ{page}: 重複データ検出。全ページ取得完了。'})}\n\n"
                        break
                    seen_keys |= page_keys

                    all_members.extend(page_members)
                    yield f"data: {json.dumps({'type': 'progress', 'current': page, 'membersSoFar': len(all_members), 'pageMembers': len(page_members)})}\n\n"

                except http_requests.exceptions.Timeout:
                    consecutive_errors += 1
                    yield f"data: {json.dumps({'type': 'warning', 'message': f'ページ {page}: タイムアウト'})}\n\n"
                    if consecutive_errors >= SUNVY_MAX_CONSECUTIVE_ERRORS:
                        yield f"data: {json.dumps({'type': 'warning', 'message': f'{SUNVY_MAX_CONSECUTIVE_ERRORS}回連続タイムアウトのため停止します。'})}\n\n"
                        break
                    time.sleep(SUNVY_PAGE_INTERVAL)
                    page += 1
                    continue
                except Exception:
                    consecutive_errors += 1
                    yield f"data: {json.dumps({'type': 'warning', 'message': f'ページ {page} の取得に失敗しました。'})}\n\n"
                    if consecutive_errors >= SUNVY_MAX_CONSECUTIVE_ERRORS:
                        yield f"data: {json.dumps({'type': 'warning', 'message': '連続エラーのため停止します。取得済みデータを保存します。'})}\n\n"
                        break
                    page += 1
                    continue

                page += 1

            # Save cache & mark fetch time
            total_pages = page - 1
            cache_data = {
                'members': all_members,
                'cachedAt': datetime.now().isoformat(),
                'totalPages': total_pages
            }
            save_json('sunvy_cache.json', cache_data)
            SUNVY_LAST_FETCH_TIME = time.time()

            yield f"data: {json.dumps({'type': 'done', 'members': all_members, 'count': len(all_members), 'cachedAt': cache_data['cachedAt'], 'totalPages': total_pages})}\n\n"

        except http_requests.exceptions.Timeout:
            yield f"data: {json.dumps({'type': 'error', 'message': 'Sunvyサーバーへの接続がタイムアウトしました'})}\n\n"
        except http_requests.exceptions.ConnectionError:
            yield f"data: {json.dumps({'type': 'error', 'message': 'Sunvyサーバーに接続できません'})}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'message': f'エラーが発生しました: {str(e)}'})}\n\n"
        finally:
            SUNVY_LOCK.release()

    return Response(stream_with_context(generate()), content_type='text/event-stream')

@app.route('/api/sunvy/cache', methods=['GET'])
@login_required
def get_sunvy_cache():
    cache = load_json('sunvy_cache.json')
    if cache.get('members'):
        return jsonify({
            'members': cache['members'],
            'count': len(cache['members']),
            'cachedAt': cache.get('cachedAt', '')
        })
    return jsonify({'members': [], 'count': 0, 'cachedAt': ''})

@app.route('/api/sunvy/cache', methods=['DELETE'])
@login_required
def clear_sunvy_cache():
    global SUNVY_LAST_FETCH_TIME
    save_json('sunvy_cache.json', {})
    SUNVY_LAST_FETCH_TIME = 0  # Reset cooldown so re-fetch is allowed
    return jsonify({'success': True})

@app.route('/api/sunvy/upload-csv', methods=['POST'])
@login_required
def upload_sunvy_csv():
    """CSVファイルからSunvyメンバーデータをインポートする"""
    if 'file' not in request.files:
        return jsonify({'error': 'ファイルがありません'}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'ファイルが選択されていません'}), 400
    if not f.filename.lower().endswith('.csv'):
        return jsonify({'error': 'CSVファイルを選択してください'}), 400

    try:
        # BOM付きUTF-8も対応
        raw = f.read()
        try:
            content = raw.decode('utf-8-sig')
        except UnicodeDecodeError:
            content = raw.decode('shift_jis', errors='replace')

        reader = csv.DictReader(io.StringIO(content))
        members = []
        skipped = 0

        for row in reader:
            nickname = (row.get('Nickname') or '').strip()
            member_id = (row.get('ID') or '').strip()
            dollars   = (row.get('9Highドル') or '').strip()
            points    = (row.get('9Highポイント') or '').strip()
            last_date = (row.get('Latest date') or '').strip()

            if not nickname or not member_id:
                skipped += 1
                continue

            # 日付を YYYY-MM-DD に正規化（時刻部分・ミリ秒を除去）
            if last_date:
                last_date = last_date.split(' ')[0].split('T')[0]

            # 0 を空文字に統一しない（そのまま保持し、表示側で判断）
            members.append({
                'name':       nickname,
                'memberCode': member_id,
                'photoUrl':   '',
                'raw':        f'{nickname}({member_id})',
                'points':     points,
                'dollars':    dollars,
                'lastVisit':  last_date,
            })

        if len(members) == 0:
            return jsonify({'error': 'CSVからメンバーデータを読み込めませんでした。列名を確認してください。'}), 400

        cache_data = {
            'members':    members,
            'cachedAt':   datetime.now().isoformat(),
            'totalPages': 0,
            'source':     'csv',
        }
        save_json('sunvy_cache.json', cache_data)

        return jsonify({
            'success':  True,
            'count':    len(members),
            'skipped':  skipped,
            'cachedAt': cache_data['cachedAt'],
        })

    except Exception as e:
        return jsonify({'error': f'CSVの解析に失敗しました: {str(e)}'}), 400


@app.route('/api/players/import-from-cache', methods=['POST'])
@login_required
def import_players_from_cache():
    """Sunvyキャッシュ（CSVまたはスクレイピング）からプレイヤーを一括インポート"""
    cache = load_json('sunvy_cache.json')
    members = cache.get('members', [])
    if not members:
        return jsonify({'error': 'キャッシュにデータがありません。先にCSVをアップロードしてください。'}), 400

    players_data = load_json('players.json')
    existing = players_data.get('players', [])

    # 既存プレイヤーのコード・名前セットを作成
    existing_codes = {p['memberCode'] for p in existing if p.get('memberCode')}
    existing_names = {(p.get('name') or '').strip() for p in existing}

    imported = 0
    skipped = 0

    for m in members:
        code = (m.get('memberCode') or '').strip()
        name = (m.get('name') or '').strip()
        if not name:
            continue

        # 重複チェック: コード優先、なければ名前
        if code and code in existing_codes:
            skipped += 1
            continue
        if not code and name in existing_names:
            skipped += 1
            continue

        new_player = {
            'id':         'p-' + uuid.uuid4().hex[:8],
            'name':       name,
            'nickname':   name,
            'memberCode': code,
            'photoUrl':   m.get('photoUrl', ''),
            'note':       '',
            'createdAt':  datetime.now().isoformat(),
        }
        existing.append(new_player)
        if code:
            existing_codes.add(code)
        existing_names.add(name)
        imported += 1

    players_data['players'] = existing
    save_json('players.json', players_data)

    return jsonify({
        'success':  True,
        'imported': imported,
        'skipped':  skipped,
        'total':    len(members),
    })


@app.route('/api/players/detect-invalid', methods=['GET'])
@admin_required
def detect_invalid_players():
    """スクレイピング由来の不正プレイヤー（数値名・日付名・極端に短い名前）を検出"""
    players_data = load_json('players.json')
    existing = players_data.get('players', [])
    # 公式CSVのIDセットを取得
    cache = load_json('sunvy_cache.json')
    valid_codes = {m.get('memberCode', '') for m in cache.get('members', []) if m.get('memberCode')}

    invalid = []
    for p in existing:
        name = (p.get('name') or '').strip()
        code = (p.get('memberCode') or '').strip()
        reason = None
        # 数値のみ（ページ番号等）
        if re.match(r'^\d+$', name):
            reason = f'名前が数値のみ「{name}」'
        # MM/DD形式（日付誤取得）
        elif re.match(r'^\d{1,2}/\d{1,2}$', name):
            reason = f'名前が日付形式「{name}」'
        # 1文字のみかつCSVに存在しないコード
        elif len(name) <= 1 and (not code or code not in valid_codes):
            reason = f'名前が1文字「{name}」かつCSV未登録'
        # コードがCSVに存在しない（CSVキャッシュがある場合のみ）
        elif valid_codes and code and code not in valid_codes:
            reason = f'IDがCSV未登録「{code}」'
        if reason:
            invalid.append({'id': p['id'], 'name': name, 'memberCode': code, 'reason': reason})

    return jsonify({'invalid': invalid, 'total': len(existing)})


@app.route('/api/players/delete-invalid', methods=['POST'])
@admin_required
def delete_invalid_players():
    """指定IDのプレイヤーを一括削除"""
    req = request.get_json()
    ids_to_delete = set(req.get('ids', []))
    if not ids_to_delete:
        return jsonify({'error': 'IDが指定されていません'}), 400

    players_data = load_json('players.json')
    before = len(players_data.get('players', []))
    players_data['players'] = [p for p in players_data.get('players', []) if p['id'] not in ids_to_delete]
    deleted = before - len(players_data['players'])
    save_json('players.json', players_data)
    return jsonify({'deleted': deleted})


@app.route('/api/players/clear-all', methods=['POST'])
@admin_required
def clear_all_players():
    """全プレイヤーデータを削除"""
    players_data = load_json('players.json')
    count = len(players_data.get('players', []))
    players_data['players'] = []
    save_json('players.json', players_data)
    return jsonify({'deleted': count})


# ==================== ROUTES: Prizes ====================
@app.route('/api/prizes', methods=['GET'])
@login_required
def get_prizes():
    return jsonify(load_prizes())

@app.route('/api/prizes/types', methods=['POST'])
@login_required
def create_prize_type():
    req = request.get_json()
    data = load_prizes()
    pt = {
        'id': 'pt-' + uuid.uuid4().hex[:8],
        'name': req.get('name', ''),
        'description': req.get('description', ''),
        'createdAt': datetime.now().isoformat()
    }
    data['types'].append(pt)
    save_json('prizes.json', data)
    return jsonify(pt), 201

@app.route('/api/prizes/types/<type_id>', methods=['DELETE'])
@login_required
def delete_prize_type(type_id):
    data = load_prizes()
    data['types'] = [t for t in data['types'] if t['id'] != type_id]
    # Also remove associated awards
    data['awards'] = [a for a in data['awards'] if a['typeId'] != type_id]
    save_json('prizes.json', data)
    return jsonify({'success': True})

@app.route('/api/prizes/awards', methods=['POST'])
@login_required
def create_award():
    req = request.get_json()
    data = load_prizes()
    today = datetime.now().strftime('%Y-%m-%d')
    award = {
        'id': 'aw-' + uuid.uuid4().hex[:8],
        'playerId':   req.get('playerId', ''),
        'typeId':     req.get('typeId', ''),
        'status':     req.get('status', '未付与'),
        'note':       req.get('note', ''),
        'occurredAt': req.get('occurredAt', today),   # 権利発生日
        'quantity':   int(req.get('quantity', 1)),     # 個数
        'createdAt':  datetime.now().isoformat(),
        'givenAt':    None
    }
    data['awards'].append(award)
    save_json('prizes.json', data)
    return jsonify(award), 201

@app.route('/api/prizes/awards/<award_id>', methods=['PUT'])
@login_required
def update_award(award_id):
    req = request.get_json()
    data = load_prizes()
    for a in data['awards']:
        if a['id'] == award_id:
            if 'status' in req:
                a['status'] = req['status']
                if req['status'] == '付与済み' and not a.get('givenAt'):
                    a['givenAt'] = datetime.now().isoformat()
                elif req['status'] == '未付与':
                    a['givenAt'] = None
            if 'note' in req:
                a['note'] = req['note']
            save_json('prizes.json', data)
            return jsonify(a)
    return jsonify({'error': 'Award not found'}), 404

@app.route('/api/prizes/awards/<award_id>', methods=['DELETE'])
@login_required
def delete_award(award_id):
    data = load_prizes()
    data['awards'] = [a for a in data['awards'] if a['id'] != award_id]
    save_json('prizes.json', data)
    return jsonify({'success': True})

# ==================== BACKUP ====================
import shutil

BACKUP_DIR = 'backups'

def create_backup(label=None):
    """data/ 以下のJSONをbackups/YYYY-MM-DD_HH-MM/にコピー"""
    ts = datetime.now().strftime('%Y-%m-%d_%H-%M')
    if label:
        ts = f"{ts}_{label}"
    dest = os.path.join(BACKUP_DIR, ts)
    os.makedirs(dest, exist_ok=True)
    copied = []
    for fname in os.listdir(DATA_DIR):
        if fname.endswith('.json'):
            shutil.copy2(os.path.join(DATA_DIR, fname), os.path.join(dest, fname))
            copied.append(fname)
    # 30世代より古いものを自動削除
    all_backups = sorted(os.listdir(BACKUP_DIR))
    for old in all_backups[:-30]:
        old_path = os.path.join(BACKUP_DIR, old)
        if os.path.isdir(old_path):
            shutil.rmtree(old_path)
    return {'timestamp': ts, 'files': copied, 'path': dest}

def backup_scheduler():
    """毎日 3:00 AM に自動バックアップ"""
    import time as _time
    while True:
        now = datetime.now()
        # 次の3:00 AMまでの秒数を計算
        next_run = now.replace(hour=3, minute=0, second=0, microsecond=0)
        if now >= next_run:
            next_run = next_run.replace(day=next_run.day + 1)
        wait_secs = (next_run - now).total_seconds()
        _time.sleep(wait_secs)
        try:
            result = create_backup('auto')
            print(f"[Backup] 自動バックアップ完了: {result['timestamp']} ({len(result['files'])}ファイル)")
        except Exception as e:
            print(f"[Backup] 自動バックアップ失敗: {e}")

@app.route('/api/backup', methods=['POST'])
@admin_required
def manual_backup():
    try:
        result = create_backup('manual')
        return jsonify({'success': True, 'timestamp': result['timestamp'], 'files': result['files']})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/backup/list', methods=['GET'])
@admin_required
def list_backups():
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        backups = []
        for name in sorted(os.listdir(BACKUP_DIR), reverse=True):
            path = os.path.join(BACKUP_DIR, name)
            if os.path.isdir(path):
                files = [f for f in os.listdir(path) if f.endswith('.json')]
                backups.append({'name': name, 'files': files})
        return jsonify({'backups': backups})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/backup/restore/<name>', methods=['POST'])
@admin_required
def restore_backup(name):
    # パストラバーサル対策
    if '..' in name or '/' in name or '\\' in name:
        return jsonify({'error': '無効なバックアップ名です'}), 400
    backup_path = os.path.join(BACKUP_DIR, name)
    if not os.path.isdir(backup_path):
        return jsonify({'error': 'バックアップが見つかりません'}), 404
    try:
        # 復元前に現在の状態を自動バックアップ
        create_backup('before_restore')
        # バックアップからdataへ上書きコピー
        restored = []
        for fname in os.listdir(backup_path):
            if fname.endswith('.json'):
                shutil.copy2(os.path.join(backup_path, fname), os.path.join(DATA_DIR, fname))
                restored.append(fname)
        return jsonify({'success': True, 'restored': restored})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== EVENT IMPORT (データ取込) ====================

def parse_pokerfans_event_page(html):
    """Parse a pokerfans.jp event detail page to extract name, date, entries, and results."""
    soup = BeautifulSoup(html, 'html.parser')

    # --- Event name: try h1 first, then og:title, then <title> ---
    name = ''
    h1 = soup.find('h1')
    if h1:
        name = h1.get_text(strip=True)
    if not name:
        og = soup.find('meta', property='og:title')
        if og:
            name = og.get('content', '').strip()
    if not name:
        title_tag = soup.find('title')
        if title_tag:
            name = title_tag.get_text(strip=True)

    # --- Date extraction (multiple strategies) ---
    date_str = ''
    full_text = soup.get_text()

    def infer_year(ev_m, ev_d):
        """Return the most likely year for a month/day with no year context."""
        now = datetime.now()
        year = now.year
        try:
            from datetime import date as _date
            event_date = _date(year, ev_m, ev_d)
            delta = (event_date - _date.today()).days
            # If the date is more than 60 days in the future it's probably last year
            if delta > 60:
                year -= 1
        except ValueError:
            pass
        return year

    # Strategy 1: <time datetime="YYYY-MM-DD...">
    for time_el in soup.find_all('time'):
        dt_attr = time_el.get('datetime', '')
        tm = re.match(r'(\d{4})-(\d{2})-(\d{2})', dt_attr)
        if tm:
            date_str = f"{tm.group(1)}-{tm.group(2)}-{tm.group(3)}"
            break

    # Strategy 2: Full YYYY年MM月DD日 / YYYY/MM/DD / YYYY-MM-DD in text
    if not date_str:
        dm = re.search(r'(\d{4})[年/\-](\d{1,2})[月/\-](\d{1,2})日?', full_text)
        if dm:
            y, m, d = dm.groups()
            date_str = f"{y}-{int(m):02d}-{int(d):02d}"

    # Strategy 3: "Start:" / "Start：" label followed by MM月DD日
    if not date_str:
        sm = re.search(r'[Ss]tart\s*[：:]\s*(\d{1,2})月(\d{1,2})日', full_text)
        if sm:
            ev_m, ev_d = int(sm.group(1)), int(sm.group(2))
            date_str = f"{infer_year(ev_m, ev_d)}-{ev_m:02d}-{ev_d:02d}"

    # Strategy 4: Any MM月DD日 pattern (first occurrence is usually the event date)
    if not date_str:
        sm2 = re.search(r'(\d{1,2})月(\d{1,2})日', full_text)
        if sm2:
            ev_m, ev_d = int(sm2.group(1)), int(sm2.group(2))
            date_str = f"{infer_year(ev_m, ev_d)}-{ev_m:02d}-{ev_d:02d}"

    # --- Entries: look for "XX/YY" near "Entries" keyword ---
    entries_current = 0
    entries_max = 0
    entries_match = re.search(r'[Ee]ntries[:\s]*(\d+)\s*/\s*(\d+)', full_text)
    if entries_match:
        entries_current = int(entries_match.group(1))
        entries_max = int(entries_match.group(2))

    # --- Tournament Result table ---
    results = []
    for table in soup.find_all('table'):
        ths = table.find_all('th')
        header_texts = [th.get_text(strip=True).lower() for th in ths]
        if any('ranking' in h or 'rank' in h for h in header_texts) or \
           any('nickname' in h or 'nick' in h or 'name' in h for h in header_texts):
            for row in table.find_all('tr'):
                cells = row.find_all('td')
                if len(cells) >= 2:
                    rank_text = cells[0].get_text(strip=True)
                    nickname_text = cells[1].get_text(strip=True)
                    points_text = cells[2].get_text(strip=True) if len(cells) > 2 else '--'
                    if rank_text and re.match(r'^\d+$', rank_text) and nickname_text:
                        results.append({
                            'ranking': int(rank_text),
                            'nickname': nickname_text,
                            'pointsRaw': points_text
                        })
            if results:
                break

    return {
        'name': name,
        'date': date_str,
        'entries': {'current': entries_current, 'max': entries_max},
        'results': results
    }


def match_nickname_to_player(nickname, players, sunvy_members, member_code=None):
    """Try to match a pokerfans/sunvy nickname to an existing player.
    Returns (player_id, match_status) where status is 'auto' or 'unmatched'.
    member_code: Sunvy member code string (digits) for direct ID matching.
    """
    if not nickname:
        return None, 'unmatched'
    nick_stripped = nickname.strip()
    if not nick_stripped or nick_stripped.lower() in ('unknown', 'unknown player', '-', ''):
        return None, 'unmatched'
    nick_lower = nick_stripped.lower()

    # 0. Direct memberCode match (most reliable — Sunvy assigns stable numeric IDs)
    if member_code:
        mc_str = str(member_code).strip()
        for p in players:
            if p.get('memberCode', '') == mc_str:
                return p['id'], 'auto'
        # Also check sunvy cache by memberCode → player cross-ref
        for m in sunvy_members:
            if str(m.get('memberCode', '')).strip() == mc_str:
                for p in players:
                    if p.get('memberCode', '') == mc_str:
                        return p['id'], 'auto'

    # 1. Exact match on player name or nickname field
    for p in players:
        if p.get('name', '').strip() == nick_stripped:
            return p['id'], 'auto'
        if p.get('nickname', '').strip() == nick_stripped:
            return p['id'], 'auto'

    # 2. Case-insensitive match
    for p in players:
        if p.get('name', '').strip().lower() == nick_lower:
            return p['id'], 'auto'
        if p.get('nickname', '') and p.get('nickname', '').strip().lower() == nick_lower:
            return p['id'], 'auto'

    # 3. Check Sunvy cache: match by name, then cross-ref by memberCode
    for m in sunvy_members:
        m_name = m.get('name', '').strip()
        if m_name == nick_stripped or m_name.lower() == nick_lower:
            mc = m.get('memberCode', '')
            if mc:
                for p in players:
                    if p.get('memberCode', '') == mc:
                        return p['id'], 'auto'

    return None, 'unmatched'


def upsert_imported_event(data, new_event):
    """Insert or update an event in data.
    Match priority:
      1. pokerfansUrl exact match (primary stable key)
      2. Same pokerfansName + same date (catches re-imports via different URL format)
    Returns (final_event_dict, is_update: bool).
    """
    events = data.setdefault('events', [])
    # Primary: URL match
    idx = next(
        (i for i, e in enumerate(events) if e.get('pokerfansUrl') == new_event['pokerfansUrl']),
        None
    )
    # Secondary: name + date match (de-duplicate across pokerfans vs applyers imports)
    if idx is None and new_event.get('pokerfansName') and new_event.get('date'):
        idx = next(
            (i for i, e in enumerate(events)
             if e.get('pokerfansName', '').strip() == new_event['pokerfansName'].strip()
             and e.get('date', '') == new_event['date']),
            None
        )

    if idx is None:
        events.append(new_event)
        return new_event, False

    old = events[idx]
    now_str = new_event['importedAt']

    # Preserve rankings that were manually matched
    manual_by_rank = {
        r['ranking']: r
        for r in old.get('results', [])
        if r.get('matchStatus') == 'manual'
    }
    for r in new_event.get('results', []):
        if r['ranking'] in manual_by_rank:
            m = manual_by_rank[r['ranking']]
            r['matchedPlayerId'] = m['matchedPlayerId']
            r['matchedPlayerName'] = m['matchedPlayerName']
            r['matchStatus'] = 'manual'

    updated = {
        **new_event,
        'importId': old['importId'],                        # keep stable ID
        'importedAt': old.get('importedAt', now_str),       # first import time
        'updatedAt': now_str,                               # freshness stamp
    }
    for key in ('addedToLeaderboard', 'addedAt', 'eventType'):
        if key in old:
            updated[key] = old[key]

    events[idx] = updated
    return updated, True


def parse_sunvy_applyers_page(html):
    """Parse a club.sunvy.jp/tours/events/XXXX/applyers page.

    The applyers page uses a two-column card layout:
      - "Players list 1/2" (left column): players in DOM top→bottom order
      - "Players list 2/2" (right column): players in DOM top→bottom order
    Ranking is assigned sequentially: all of col-1 first, then all of col-2.
    Rank badges like "1st Rank", "2Rank" etc. are IGNORED — not used for ordering.

    Player format in card text: NAME(ID:MEMBERCODE) => Prepaid:¥0 => ...

    Returns {'name': str, 'date': str (YYYY-MM-DD),
             'results': [{'ranking': int, 'nickname': str, 'memberCode': str, 'pointsRaw': str}]}
    """
    soup = BeautifulSoup(html, 'html.parser')

    name = ''
    date_str = ''
    info_tables = set()

    # --- Step 1: Extract tournament name and date from info table ---
    for table in soup.find_all('table'):
        rows = table.find_all('tr')
        for row in rows:
            cells = row.find_all(['th', 'td'])
            texts = [c.get_text(strip=True) for c in cells]
            if len(texts) < 2:
                continue
            label = texts[0].strip().lower()
            value = texts[1].strip()
            if 'tournament name' in label or 'トーナメント名' in label:
                name = re.sub(r'\s*\(Processed\)\s*', '', value, flags=re.IGNORECASE).strip()
                info_tables.add(id(table))
            if 'match date' in label or '開催日' in label or '試合日' in label:
                raw_date = value
                dm = re.search(r'(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})', raw_date)
                if dm:
                    date_str = f"{dm.group(1)}-{int(dm.group(2)):02d}-{int(dm.group(3)):02d}"
                info_tables.add(id(table))

    # Fallback date from page text
    if not date_str:
        full_text = soup.get_text()
        dm = re.search(r'(\d{4})[年/\-](\d{1,2})[月/\-](\d{1,2})日?', full_text)
        if dm:
            date_str = f"{dm.group(1)}-{int(dm.group(2)):02d}-{int(dm.group(3)):02d}"

    # --- Step 2: Extract players from two-column card layout ---
    # Regex to find player entries: NAME(ID:MEMBERCODE)
    # Rank badge prefixes to strip: "1Rank", "1st Rank", "2nd Rank", "3rd Rank", "4th Rank" etc.
    PLAYER_PAT = re.compile(r'(.+?)\(ID:(\d+)\)')
    RANK_BADGE_PAT = re.compile(r'^\d+(?:st|nd|rd|th)?\s*Rank\s*', re.IGNORECASE)

    def extract_players_from_container(container):
        """Extract player (name, memberCode) pairs from a DOM container, in DOM order."""
        found = []
        seen_codes = set()
        # Walk all text nodes and elements that contain the ID: pattern
        for el in container.find_all(string=PLAYER_PAT):
            text = el.strip()
            m = PLAYER_PAT.search(text)
            if not m:
                continue
            raw_name = m.group(1).strip()
            member_code = m.group(2).strip()
            if member_code in seen_codes:
                continue
            # Strip any rank badge prefix from the name
            clean_name = RANK_BADGE_PAT.sub('', raw_name).strip()
            if clean_name:
                found.append({'nickname': clean_name, 'memberCode': member_code})
                seen_codes.add(member_code)
        return found

    players_ordered = []

    # Find column headers "Players list 1/2" and "Players list 2/2" in DOM order
    col_header_els = []
    for el in soup.find_all(string=re.compile(r'Players\s+list\s+\d+/\d+', re.IGNORECASE)):
        col_header_els.append(el)

    if col_header_els:
        seen_member_codes = set()
        for header_text_node in col_header_els:
            # Walk up to find a column container ancestor (div/section/td)
            # that is likely to contain all the player cards for this column
            container = header_text_node.parent
            # Walk up until we find a container with meaningful depth
            for _ in range(6):
                parent = container.parent
                if parent is None or parent.name in ('body', 'html', '[document]'):
                    break
                # If the parent also contains the header text, keep going up
                if parent.find(string=re.compile(r'Players\s+list\s+\d+/\d+', re.IGNORECASE)):
                    # But stop if the parent contains BOTH column headers —
                    # that means we've gone too far up (we'd get all players from both cols)
                    header_count = len(parent.find_all(
                        string=re.compile(r'Players\s+list\s+\d+/\d+', re.IGNORECASE)
                    ))
                    if header_count > 1:
                        break  # this ancestor spans both columns — stop here
                    container = parent
                else:
                    break

            col_players = extract_players_from_container(container)
            for p in col_players:
                if p['memberCode'] not in seen_member_codes:
                    players_ordered.append(p)
                    seen_member_codes.add(p['memberCode'])
    else:
        # Fallback: no "Players list X/Y" headers found.
        # Scan the whole page in DOM order, deduplicate by memberCode.
        seen_member_codes = set()
        for el in soup.find_all(string=PLAYER_PAT):
            text = el.strip()
            m = PLAYER_PAT.search(text)
            if not m:
                continue
            raw_name = m.group(1).strip()
            member_code = m.group(2).strip()
            if member_code in seen_member_codes:
                continue
            clean_name = RANK_BADGE_PAT.sub('', raw_name).strip()
            if clean_name:
                players_ordered.append({'nickname': clean_name, 'memberCode': member_code})
                seen_member_codes.add(member_code)

    results = [
        {'ranking': i + 1, 'nickname': p['nickname'], 'memberCode': p['memberCode'], 'pointsRaw': '--'}
        for i, p in enumerate(players_ordered)
    ]
    return {'name': name, 'date': date_str, 'results': results}


@app.route('/api/sunvy/import-status', methods=['GET'])
@login_required
def get_import_status():
    """Return whether an import is currently running and the recent log."""
    running = SUNVY_EVENT_LOCK.locked()
    with _import_log_lock:
        recent = list(_import_log[-50:])
    return jsonify({'running': running, 'recentLog': recent})


@app.route('/api/sunvy/import-cancel', methods=['POST'])
@login_required
def cancel_import():
    """Force-release stuck import lock (emergency use only)."""
    if SUNVY_EVENT_LOCK.locked():
        try:
            SUNVY_EVENT_LOCK.release()
            _append_import_log('status', '手動でロックを解除しました')
            return jsonify({'success': True, 'message': 'ロックを解除しました'})
        except RuntimeError as e:
            return jsonify({'success': False, 'error': str(e)})
    return jsonify({'success': False, 'error': 'ロックは保持されていません'})


def _logged_sse_generator(gen):
    """Wrap an SSE generator: mirror every yielded message to _import_log."""
    for chunk in gen:
        try:
            payload = json.loads(chunk[len('data: '):].strip())
            _append_import_log(payload.get('type', 'status'), payload.get('message', ''))
        except Exception:
            pass
        yield chunk


@app.route('/api/sunvy/import-events', methods=['POST'])
@login_required
def import_sunvy_events():
    """SSE: login to Sunvy → paginate events list → follow applyers links →
    parse tournament name & player list → match players → stream progress → save."""
    if not SUNVY_EVENT_LOCK.acquire(blocking=False):
        return jsonify({'error': '別のイベント取込処理が実行中です。完了までお待ちください。'}), 409

    req = request.get_json()
    # Credentials: prefer request body, fall back to saved config
    config = load_config()
    sc = config.get('sunvy_creds', {})
    email = (req.get('email', '') or sc.get('email', '')).strip()
    password = (req.get('password', '') or sc.get('password', '')).strip()
    mode = req.get('mode', 'single')
    single_date = req.get('date', '').strip()
    start_date = req.get('startDate', '').strip()
    end_date = req.get('endDate', '').strip()
    if mode == 'single':
        eff_start = single_date
        eff_end = single_date
    else:
        eff_start = start_date
        eff_end = end_date

    if not email or not password:
        SUNVY_EVENT_LOCK.release()
        return jsonify({'error': 'Sunvyのメールアドレスとパスワードを設定してください（⚙ボタンから）'}), 400

    def generate():
        try:
            yield f"data: {json.dumps({'type': 'status', 'message': 'Sunvyにログイン中...'})}\n\n"
            try:
                s = sunvy_login(email, password)
            except Exception as login_err:
                yield f"data: {json.dumps({'type': 'error', 'message': str(login_err)})}\n\n"
                return
            yield f"data: {json.dumps({'type': 'status', 'message': 'ログイン成功'})}\n\n"

            players_data = load_json('players.json')
            players = players_data.get('players', [])
            sunvy_cache = load_json('sunvy_cache.json')
            sunvy_members = sunvy_cache.get('members', [])

            yield f"data: {json.dumps({'type': 'status', 'message': 'Sunvyイベント一覧を取得中...'})}\n\n"

            # --- Phase 1: collect applyers URLs ---
            applyers_links = []  # list of (url, event_id)

            def build_events_url(page_num, use_date_filter=True):
                params = f'size=10&page={page_num}'
                if use_date_filter:
                    if eff_start:
                        params = f'startDate={urllib.parse.quote(eff_start)}&' + params
                    if eff_end:
                        params += f'&endDate={urllib.parse.quote(eff_end)}'
                return f'https://club.sunvy.jp/tours/events/all?{params}'

            date_filter_ok = True

            for page in range(0, SUNVY_MAX_PAGES):
                events_url = build_events_url(page, use_date_filter=date_filter_ok)
                yield f"data: {json.dumps({'type': 'progress', 'message': f'URL: {events_url}'})}\n\n"
                try:
                    resp = s.get(events_url, timeout=SUNVY_REQUEST_TIMEOUT)
                    if '/login' in resp.url:
                        yield f"data: {json.dumps({'type': 'status', 'message': 'セッション切れ。自動再ログイン中...'})}\n\n"
                        try:
                            s = sunvy_login(email, password)
                            resp = s.get(events_url, timeout=SUNVY_REQUEST_TIMEOUT)
                            yield f"data: {json.dumps({'type': 'status', 'message': '再ログイン成功。処理を継続します'})}\n\n"
                        except Exception as relogin_err:
                            yield f"data: {json.dumps({'type': 'error', 'message': f'再ログイン失敗: {str(relogin_err)}'})}\n\n"
                            return
                        if '/login' in resp.url:
                            yield f"data: {json.dumps({'type': 'error', 'message': '再ログインに失敗しました'})}\n\n"
                            return
                    if resp.status_code != 200:
                        body_snippet = resp.text[:400].replace('\n', ' ').strip()
                        if resp.status_code == 400 and date_filter_ok and (eff_start or eff_end):
                            date_filter_ok = False
                            events_url = build_events_url(page, use_date_filter=False)
                            resp = s.get(events_url, timeout=SUNVY_REQUEST_TIMEOUT)
                            if resp.status_code != 200:
                                yield f"data: {json.dumps({'type': 'warning', 'message': f'再試行も失敗: HTTP {resp.status_code}'})}\n\n"
                                break
                        else:
                            yield f"data: {json.dumps({'type': 'warning', 'message': f'ページ{page}: HTTP {resp.status_code}'})}\n\n"
                            break

                    soup = BeautifulSoup(resp.text, 'html.parser')
                    page_text = soup.get_text()
                    if 'データがありません' in page_text or 'No Data' in page_text:
                        yield f"data: {json.dumps({'type': 'status', 'message': f'ページ{page}: データなし。全ページ完了。'})}\n\n"
                        break

                    # Build a fast lookup set of already-collected URLs for O(1) dedup
                    collected_urls = {u for u, _ in applyers_links}

                    # Collect applyers links WITH per-event date extraction
                    # so we can filter in Phase 1 instead of fetching 700+ irrelevant pages.
                    page_links = []
                    seen_on_page = set()       # within-page dedup
                    page_event_dates = []      # dates of events found on this page (for stop check)
                    all_before_start = True    # used for early-stop detection

                    def extract_link_date(anchor_el):
                        """Walk up the DOM from a link to find a nearby event date (YYYY-MM-DD).
                        Stops at the first ancestor whose text is short enough to be a single card."""
                        node = anchor_el.parent
                        for _ in range(6):
                            if node is None:
                                break
                            node_text = node.get_text(' ', strip=True)
                            # Only look inside containers small enough to belong to one event
                            if len(node_text) < 600:
                                dm = re.search(r'(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})', node_text)
                                if dm:
                                    y, mo, d = dm.group(1), dm.group(2), dm.group(3)
                                    if 2020 <= int(y) <= 2035:
                                        return f"{y}-{int(mo):02d}-{int(d):02d}"
                            node = node.parent
                        return ''

                    for a in soup.find_all('a', href=True):
                        href = a['href'].strip()
                        am = re.search(r'/tours/events/(\d+)/applyers', href)
                        if not am:
                            continue
                        event_id = am.group(1)
                        full_url = f'https://club.sunvy.jp/tours/events/{event_id}/applyers'

                        # Dedup: skip URLs already collected globally or earlier on this page
                        if full_url in collected_urls or full_url in seen_on_page:
                            continue
                        seen_on_page.add(full_url)

                        # Extract the event's date from surrounding HTML
                        link_date = extract_link_date(a)
                        if link_date:
                            page_event_dates.append(link_date)

                        # Per-event date-range filter (only when we have a date AND mode=range)
                        if mode == 'range' and link_date:
                            if eff_end and link_date > eff_end:
                                continue  # event is after end date — skip (but don't stop yet)
                            if eff_start and link_date < eff_start:
                                continue  # event is before start date — skip
                            all_before_start = False

                        page_links.append((full_url, event_id))

                    if not page_links and page == 0 and not applyers_links:
                        yield f"data: {json.dumps({'type': 'status', 'message': 'applyers リンクが見つかりませんでした。イベントが存在するか確認してください。'})}\n\n"
                        break

                    if page_links:
                        applyers_links.extend(page_links)
                        in_range = len(page_links)
                        total_on_page = len(seen_on_page)
                        skipped = total_on_page - in_range
                        skip_note = f'（{skipped}件スキップ）' if skipped else ''
                        yield f"data: {json.dumps({'type': 'progress', 'message': f'ページ{page}: {in_range}件追加{skip_note}（合計{len(applyers_links)}件）'})}\n\n"
                    else:
                        yield f"data: {json.dumps({'type': 'progress', 'message': f'ページ{page}: 対象イベントなし（スキップ）'})}\n\n"

                    if mode == 'single':
                        break

                    # Early-stop for range mode:
                    # Stop when all events on the page (with dates) are before the start date
                    if mode == 'range' and eff_start and page_event_dates:
                        if max(page_event_dates) < eff_start:
                            yield f"data: {json.dumps({'type': 'status', 'message': f'ページ{page}の最新日付 {max(page_event_dates)} が開始日 {eff_start} より前のため終了'})}\n\n"
                            break

                except http_requests.exceptions.Timeout:
                    yield f"data: {json.dumps({'type': 'warning', 'message': f'ページ{page}: タイムアウト'})}\n\n"
                    break
                except Exception as ex:
                    yield f"data: {json.dumps({'type': 'warning', 'message': f'ページ{page}エラー: {str(ex)}'})}\n\n"
                    break

                wait_sec = random.randint(3, 7)
                time.sleep(wait_sec)

            if not applyers_links:
                yield f"data: {json.dumps({'type': 'done', 'events': [], 'count': 0, 'message': '取得対象のイベントが見つかりませんでした'})}\n\n"
                return

            yield f"data: {json.dumps({'type': 'status', 'message': f'{len(applyers_links)}件のイベントページを取得します...'})}\n\n"

            # --- Phase 2: fetch each applyers page (with pagination) ---
            new_events = []
            for i, (applyers_url, event_id) in enumerate(applyers_links):
                if i > 0:
                    wait_sec = random.randint(3, 7)
                    yield f"data: {json.dumps({'type': 'waiting', 'message': f'{wait_sec}秒待機中... ({i+1}/{len(applyers_links)})'})}\n\n"
                    time.sleep(wait_sec)

                yield f"data: {json.dumps({'type': 'status', 'message': f'取得中 ({i+1}/{len(applyers_links)}): {applyers_url}'})}\n\n"

                try:
                    all_players = []
                    event_name = ''
                    event_date = ''
                    seen_nicks = set()
                    page_num = 0

                    while True:
                        paged_url = applyers_url if page_num == 0 else f'{applyers_url}?page={page_num}'
                        ap_resp = s.get(paged_url, timeout=SUNVY_REQUEST_TIMEOUT)
                        if not ap_resp.ok:
                            yield f"data: {json.dumps({'type': 'warning', 'message': f'HTTP {ap_resp.status_code}: {paged_url}'})}\n\n"
                            break

                        parsed = parse_sunvy_applyers_page(ap_resp.text)

                        if not event_name and parsed.get('name'):
                            event_name = parsed['name']
                        if not event_date and parsed.get('date'):
                            event_date = parsed['date']

                        new_players = [r for r in parsed['results'] if r['nickname'] not in seen_nicks]
                        if not new_players:
                            break  # no new players = last page
                        for r in new_players:
                            seen_nicks.add(r['nickname'])
                        all_players.extend(new_players)

                        # Check if there's a next page via pagination
                        ap_soup = BeautifulSoup(ap_resp.text, 'html.parser')
                        pagination = ap_soup.find('ul', class_='pagination')
                        has_next = False
                        if pagination:
                            for a in pagination.find_all('a'):
                                href = a.get('href', '')
                                pm = re.search(r'[?&]page=(\d+)', href)
                                if pm and int(pm.group(1)) > page_num:
                                    has_next = True
                                    break
                        if not has_next:
                            break
                        page_num += 1
                        time.sleep(2)

                    # Re-assign sequential rankings starting from 1
                    all_results_raw = [
                        {
                            'ranking': idx + 1,
                            'nickname': r['nickname'],
                            'memberCode': r.get('memberCode', ''),
                            'pointsRaw': '--'
                        }
                        for idx, r in enumerate(all_players)
                    ]

                    # --- Date-range filter ---
                    if event_date and mode == 'range':
                        if eff_end and event_date > eff_end:
                            yield f"data: {json.dumps({'type': 'status', 'message': f'スキップ: {event_date} (終了日 {eff_end} より後)'})}\n\n"
                            continue
                        if eff_start and event_date < eff_start:
                            yield f"data: {json.dumps({'type': 'status', 'message': f'{event_date} が開始日 {eff_start} より前のため取込を終了します'})}\n\n"
                            break

                    # Match players
                    matched_results = []
                    for r in all_results_raw:
                        pid, status = match_nickname_to_player(
                            r['nickname'], players, sunvy_members,
                            member_code=r.get('memberCode', '')
                        )
                        matched_name = None
                        if pid:
                            mp = next((p for p in players if p['id'] == pid), None)
                            matched_name = mp['name'] if mp else None
                        matched_results.append({
                            'ranking': r['ranking'],
                            'nickname': r['nickname'],
                            'memberCode': r.get('memberCode', ''),
                            'pointsRaw': r['pointsRaw'],
                            'matchedPlayerId': pid,
                            'matchedPlayerName': matched_name,
                            'matchStatus': status
                        })

                    candidate = {
                        'importId': f"ev-{uuid.uuid4().hex[:8]}",
                        'pokerfansUrl': applyers_url,  # reuse field as stable event key
                        'pokerfansName': event_name or f'イベント {event_id}',
                        'date': event_date,
                        'entries': {'current': len(matched_results), 'max': 0},
                        'results': matched_results,
                        'importedAt': datetime.now().isoformat(),
                        'addedToLeaderboard': False
                    }

                    imported_data = load_json('sunvy_imported_events.json')
                    event, is_update = upsert_imported_event(imported_data, candidate)
                    imported_data['lastImportAt'] = datetime.now().isoformat()
                    save_json('sunvy_imported_events.json', imported_data)
                    new_events.append((event, is_update))

                    auto_count = sum(1 for r in matched_results if r['matchStatus'] == 'auto')
                    action = '更新' if is_update else '新規'
                    yield f"data: {json.dumps({'type': 'event', 'event': event, 'matched': auto_count, 'total': len(matched_results), 'action': action})}\n\n"

                except http_requests.exceptions.Timeout:
                    yield f"data: {json.dumps({'type': 'warning', 'message': f'タイムアウト: {applyers_url}'})}\n\n"
                    continue
                except Exception as ex:
                    import traceback
                    yield f"data: {json.dumps({'type': 'warning', 'message': f'解析エラー ({applyers_url}): {str(ex)}'})}\n\n"
                    continue

            added   = sum(1 for _, upd in new_events if not upd)
            updated = sum(1 for _, upd in new_events if upd)
            events_only = [ev for ev, _ in new_events]
            parts = []
            if added:   parts.append(f'{added}件新規取込')
            if updated: parts.append(f'{updated}件更新')
            msg = '、'.join(parts) if parts else '新規イベントはありませんでした'
            yield f"data: {json.dumps({'type': 'done', 'events': events_only, 'count': added, 'message': msg})}\n\n"

        except http_requests.exceptions.ConnectionError:
            yield f"data: {json.dumps({'type': 'error', 'message': 'サーバーに接続できません'})}\n\n"
        except Exception as ex:
            yield f"data: {json.dumps({'type': 'error', 'message': str(ex)})}\n\n"
        finally:
            SUNVY_EVENT_LOCK.release()

    _clear_import_log()
    return Response(stream_with_context(_logged_sse_generator(generate())), content_type='text/event-stream')


@app.route('/api/sunvy/imported-events', methods=['GET'])
@login_required
def get_imported_events():
    """Return all imported events."""
    data = load_json('sunvy_imported_events.json')
    return jsonify(data)


@app.route('/api/sunvy/imported-events/backfill-dates', methods=['POST'])
@login_required
def backfill_event_dates():
    """Re-fetch pokerfans pages for events whose date is missing and update in place."""
    data = load_json('sunvy_imported_events.json')
    events = data.get('events', [])
    fixed = 0
    errors = []
    for ev in events:
        if ev.get('date'):
            continue
        url = ev.get('pokerfansUrl', '')
        if not url:
            continue
        try:
            resp = http_requests.get(url, timeout=SUNVY_REQUEST_TIMEOUT,
                headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})
            if resp.status_code == 200:
                parsed = parse_pokerfans_event_page(resp.text)
                if parsed.get('date'):
                    ev['date'] = parsed['date']
                    fixed += 1
            time.sleep(2)
        except Exception as ex:
            errors.append(str(ex))
    save_json('sunvy_imported_events.json', data)
    return jsonify({'fixed': fixed, 'errors': errors})


@app.route('/api/sunvy/imported-events/<import_id>/match', methods=['POST'])
@login_required
def update_event_match(import_id):
    """Manually update the player match for one ranking entry in an imported event."""
    req = request.get_json()
    ranking = req.get('ranking')
    player_id = req.get('playerId')  # None = unmatch

    data = load_json('sunvy_imported_events.json')
    for event in data.get('events', []):
        if event.get('importId') != import_id:
            continue
        for r in event.get('results', []):
            if r.get('ranking') != ranking:
                continue
            if player_id:
                players_data = load_json('players.json')
                player = next((p for p in players_data.get('players', []) if p['id'] == player_id), None)
                r['matchedPlayerId'] = player_id
                r['matchedPlayerName'] = player['name'] if player else None
                r['matchStatus'] = 'manual'
            else:
                r['matchedPlayerId'] = None
                r['matchedPlayerName'] = None
                r['matchStatus'] = 'unmatched'
            save_json('sunvy_imported_events.json', data)
            return jsonify({'success': True, 'result': r})
    return jsonify({'error': 'Not found'}), 404


@app.route('/api/sunvy/import-event-direct', methods=['POST'])
@login_required
def import_event_direct():
    """SSE: fetch a single club.sunvy.jp/tours/events/XXXX/applyers URL directly."""
    if not SUNVY_EVENT_LOCK.acquire(blocking=False):
        return jsonify({'error': '別のイベント取込処理が実行中です。完了までお待ちください。'}), 409

    req = request.get_json()
    url = req.get('url', '').strip()

    if not url:
        SUNVY_EVENT_LOCK.release()
        return jsonify({'error': 'URLを入力してください'}), 400

    # Normalise: accept the event detail page and auto-append /applyers
    m_event = re.search(r'club\.sunvy\.jp/tours/events/(\d+)(/applyers)?', url)
    if not m_event:
        SUNVY_EVENT_LOCK.release()
        return jsonify({'error': 'club.sunvy.jp/tours/events/.../applyers のURLを入力してください'}), 400

    event_id = m_event.group(1)
    applyers_url = f'https://club.sunvy.jp/tours/events/{event_id}/applyers'

    # Load saved Sunvy credentials
    config = load_config()
    sc = config.get('sunvy_creds', {})
    email = sc.get('email', '').strip()
    password = sc.get('password', '').strip()
    if not email or not password:
        SUNVY_EVENT_LOCK.release()
        return jsonify({'error': 'Sunvyの認証情報が設定されていません（⚙ボタンから設定してください）'}), 400

    def generate():
        try:
            yield f"data: {json.dumps({'type': 'status', 'message': 'Sunvyにログイン中...'})}\n\n"
            try:
                s = sunvy_login(email, password)
            except Exception as login_err:
                yield f"data: {json.dumps({'type': 'error', 'message': str(login_err)})}\n\n"
                return
            yield f"data: {json.dumps({'type': 'status', 'message': 'ログイン成功'})}\n\n"

            players_data = load_json('players.json')
            players = players_data.get('players', [])
            sunvy_cache = load_json('sunvy_cache.json')
            sunvy_members = sunvy_cache.get('members', [])

            try:
                all_players = []
                event_name = ''
                event_date = ''
                seen_nicks = set()
                page_num = 0

                while True:
                    paged_url = applyers_url if page_num == 0 else f'{applyers_url}?page={page_num}'
                    yield f"data: {json.dumps({'type': 'status', 'message': f'取得中: {paged_url}'})}\n\n"
                    ap_resp = s.get(paged_url, timeout=SUNVY_REQUEST_TIMEOUT)
                    if not ap_resp.ok:
                        yield f"data: {json.dumps({'type': 'error', 'message': f'HTTP {ap_resp.status_code}: {paged_url}'})}\n\n"
                        return

                    parsed = parse_sunvy_applyers_page(ap_resp.text)

                    if not event_name and parsed.get('name'):
                        event_name = parsed['name']
                    if not event_date and parsed.get('date'):
                        event_date = parsed['date']

                    new_players = [r for r in parsed['results'] if r['nickname'] not in seen_nicks]
                    if not new_players:
                        break
                    for r in new_players:
                        seen_nicks.add(r['nickname'])
                    all_players.extend(new_players)

                    ap_soup = BeautifulSoup(ap_resp.text, 'html.parser')
                    pagination = ap_soup.find('ul', class_='pagination')
                    has_next = False
                    if pagination:
                        for a in pagination.find_all('a'):
                            href = a.get('href', '')
                            pm = re.search(r'[?&]page=(\d+)', href)
                            if pm and int(pm.group(1)) > page_num:
                                has_next = True
                                break
                    if not has_next:
                        break
                    page_num += 1
                    time.sleep(2)

                yield f"data: {json.dumps({'type': 'status', 'message': f'{len(all_players)}名を取得しました'})}\n\n"

                # Re-assign sequential rankings
                all_results_raw = [
                    {
                        'ranking': idx + 1,
                        'nickname': r['nickname'],
                        'memberCode': r.get('memberCode', ''),
                        'pointsRaw': '--'
                    }
                    for idx, r in enumerate(all_players)
                ]

                matched_results = []
                for r in all_results_raw:
                    pid, status = match_nickname_to_player(
                        r['nickname'], players, sunvy_members,
                        member_code=r.get('memberCode', '')
                    )
                    matched_name = None
                    if pid:
                        mp = next((p for p in players if p['id'] == pid), None)
                        matched_name = mp['name'] if mp else None
                    matched_results.append({
                        'ranking': r['ranking'],
                        'nickname': r['nickname'],
                        'memberCode': r.get('memberCode', ''),
                        'pointsRaw': r['pointsRaw'],
                        'matchedPlayerId': pid,
                        'matchedPlayerName': matched_name,
                        'matchStatus': status
                    })

                candidate = {
                    'importId': f"ev-{uuid.uuid4().hex[:8]}",
                    'pokerfansUrl': applyers_url,
                    'pokerfansName': event_name or f'イベント {event_id}',
                    'date': event_date,
                    'entries': {'current': len(matched_results), 'max': 0},
                    'results': matched_results,
                    'importedAt': datetime.now().isoformat(),
                    'addedToLeaderboard': False
                }

                imported_data = load_json('sunvy_imported_events.json')
                event, is_update = upsert_imported_event(imported_data, candidate)
                imported_data['lastImportAt'] = datetime.now().isoformat()
                save_json('sunvy_imported_events.json', imported_data)

                auto_count = sum(1 for r in matched_results if r['matchStatus'] == 'auto')
                action = '更新' if is_update else '新規'
                yield f"data: {json.dumps({'type': 'event', 'event': event, 'matched': auto_count, 'total': len(matched_results), 'action': action})}\n\n"
                ev_name = event_name or f'イベント {event_id}'
                verb = '更新しました' if is_update else '取込みました'
                yield f"data: {json.dumps({'type': 'done', 'events': [event], 'count': 1, 'message': f'{ev_name} を{verb}'})}\n\n"

            except http_requests.exceptions.Timeout:
                yield f"data: {json.dumps({'type': 'error', 'message': f'タイムアウト: {applyers_url}'})}\n\n"
            except Exception as ex:
                yield f"data: {json.dumps({'type': 'error', 'message': f'解析エラー: {str(ex)}'})}\n\n"

        except Exception as ex:
            yield f"data: {json.dumps({'type': 'error', 'message': f'エラー: {str(ex)}'})}\n\n"
        finally:
            SUNVY_EVENT_LOCK.release()

    _clear_import_log()
    return Response(stream_with_context(_logged_sse_generator(generate())), content_type='text/event-stream')


@app.route('/api/sunvy/imported-events/categorize', methods=['POST'])
@login_required
def categorize_imported_events():
    """Set eventType (ring/tournament) on imported events; delete non-import (skip) events."""
    req = request.get_json()
    categories = req.get('categories', [])  # [{importId, category: 'ring'|'tournament'|'skip'}]

    data = load_json('sunvy_imported_events.json')
    events = data.get('events', [])

    skip_ids = {c['importId'] for c in categories if c['category'] == 'skip'}
    cat_map = {c['importId']: c['category'] for c in categories if c['category'] != 'skip'}

    events = [e for e in events if e.get('importId') not in skip_ids]
    for e in events:
        if e.get('importId') in cat_map:
            e['eventType'] = cat_map[e['importId']]

    data['events'] = events
    save_json('sunvy_imported_events.json', data)
    return jsonify({'success': True, 'removed': len(skip_ids)})


@app.route('/api/sunvy/imported-events/<import_id>', methods=['DELETE'])
@login_required
def delete_imported_event(import_id):
    """Delete a single imported event."""
    data = load_json('sunvy_imported_events.json')
    before = len(data.get('events', []))
    data['events'] = [e for e in data.get('events', []) if e.get('importId') != import_id]
    if len(data['events']) == before:
        return jsonify({'error': 'Not found'}), 404
    save_json('sunvy_imported_events.json', data)
    return jsonify({'success': True})


@app.route('/api/sunvy/imported-events/<import_id>/add-to-leaderboard', methods=['POST'])
@login_required
def add_event_to_leaderboard(import_id):
    """Finalize an imported event and upsert it into tournaments.json.
    Re-adding is allowed (e.g. after the entry was deleted from tournament results).
    Uses pokerfansUrl as the stable key for upsert so no duplicates are created.
    """
    data = load_json('sunvy_imported_events.json')
    event = next((e for e in data.get('events', []) if e.get('importId') == import_id), None)
    if not event:
        return jsonify({'error': 'イベントが見つかりません'}), 404

    # playerOverrides: [{ranking, rank?, points?}] from the frontend editor
    req = request.get_json(silent=True) or {}
    overrides = {o['ranking']: o for o in req.get('playerOverrides', [])}

    # Build results — only include matched players
    results = []
    for r in event.get('results', []):
        if r.get('matchedPlayerId'):
            ov = overrides.get(r['ranking'], {})
            participation = ov.get('participation', False)
            final_rank   = None if participation else ov.get('rank', r['ranking'])
            final_points = ov.get('points', 0)
            results.append({
                'playerId': r['matchedPlayerId'],
                'playerName': r.get('matchedPlayerName', ''),
                'rank': final_rank,
                'points': final_points,
                'placed': not participation
            })

    tournaments_data = load_json('tournaments.json')
    if 'results' not in tournaments_data:
        tournaments_data['results'] = []

    event_name = re.sub(r'\s*\(Processed\)\s*', '', event.get('pokerfansName', ''), flags=re.IGNORECASE).strip()
    pokerfans_url = event.get('pokerfansUrl', '')

    new_result = {
        'tournamentId': event_name,       # event name shown in tournament results table
        'date': event.get('date', ''),
        'entries': event.get('entries', {}).get('current', len(results)),
        'name': event_name,
        'pokerfansUrl': pokerfans_url,
        'results': results
    }

    # Upsert by pokerfansUrl — replace existing entry if found, otherwise append
    existing_idx = next(
        (i for i, r in enumerate(tournaments_data['results']) if r.get('pokerfansUrl') == pokerfans_url),
        None
    )
    if existing_idx is not None:
        tournaments_data['results'][existing_idx] = new_result
    else:
        tournaments_data['results'].append(new_result)
    save_json('tournaments.json', tournaments_data)

    event['addedToLeaderboard'] = True
    event['addedAt'] = datetime.now().isoformat()
    save_json('sunvy_imported_events.json', data)

    return jsonify({'success': True, 'entry': new_result})


@app.route('/api/sunvy/imported-events/<import_id>/add-to-ringgame', methods=['POST'])
@login_required
def add_event_to_ringgame(import_id):
    """Add an imported ring-game event as a session group.
    Deduplicates by pokerfansUrl — re-adding overwrites the previous session from the same event.
    """
    data = load_json('sunvy_imported_events.json')
    event = next((e for e in data.get('events', []) if e.get('importId') == import_id), None)
    if not event:
        return jsonify({'error': 'イベントが見つかりません'}), 404

    pokerfans_url = event.get('pokerfansUrl', '') or None
    ring_data = load_ringgames()

    # Remove any session previously added from this event (dedup by pokerfansUrl)
    if pokerfans_url:
        ring_data['sessions'] = [s for s in ring_data['sessions']
                                  if s.get('pokerfansUrl') != pokerfans_url]

    # playerOverrides: [{ranking, netChips}] from the frontend editor
    req = request.get_json(silent=True) or {}
    overrides = {o['ranking']: o for o in req.get('playerOverrides', [])}

    # Build player list — all players, matched or not
    players = []
    for r in event.get('results', []):
        player_id = r.get('matchedPlayerId')
        player_name = (r.get('matchedPlayerName') or r.get('nickname', ''))
        ov = overrides.get(r['ranking'], {})
        players.append({
            'playerId': player_id or None,
            'playerName': player_name,
            'netChips': ov.get('netChips', 0)
        })

    event_date = event.get('date', '')
    raw_name = event.get('pokerfansName', '')
    event_name = re.sub(r'\s*\(Processed\)\s*', '', raw_name, flags=re.IGNORECASE).strip() or (event_date + ' セッション')
    new_session = {
        'id': str(uuid.uuid4()),
        'name': event_name,
        'date': event_date,
        'pokerfansUrl': pokerfans_url,
        'players': players
    }
    ring_data['sessions'].append(new_session)
    save_json('ringgames.json', ring_data)

    event['addedToLeaderboard'] = True
    event['addedAt'] = datetime.now().isoformat()
    save_json('sunvy_imported_events.json', data)

    return jsonify({'success': True, 'added': len(players), 'sessionId': new_session['id']})


# ==================== RUN ====================
if __name__ == '__main__':
    load_config()
    load_players()
    # バックアップスケジューラーをバックグラウンドで起動（debug=Trueのリローダー対策）
    if not os.environ.get('WERKZEUG_RUN_MAIN'):
        _t = threading.Thread(target=backup_scheduler, daemon=True)
        _t.start()
        print("[Backup] 自動バックアップスケジューラー起動（毎日 3:00 AM）")
    app.run(host='0.0.0.0', port=5000, debug=True)
